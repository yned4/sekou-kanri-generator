"""3工事の差分分析スクリプト: アプリ出力 vs 目標output"""
import os, sys, re, unicodedata
import pandas as pd
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from extractor import (
    extract_suryo, build_match_detail, filter_by_row_labels,
    get_implicit_hinshitsu_labels, build_suryo_match_map, SURYO_LEVEL_COLS,
)
from excel_writer import write_excel
import openpyxl


def norm(s):
    if not s or s == "None":
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKC", s)
    s = re.sub(r"\s+", "", s)
    return s


# ── DB読み込み ──
DB_PATH = Path(__file__).resolve().parent.parent / "database" / "kojyo_kijun.xlsx"
db_d = pd.read_excel(str(DB_PATH), sheet_name="出来形管理", dtype=str).fillna("")
db_h = pd.read_excel(str(DB_PATH), sheet_name="品質管理", dtype=str).fillna("")
db_p = pd.read_excel(str(DB_PATH), sheet_name="撮影箇所", dtype=str).fillna("")

kojyo_data = {"出来形管理": db_d.copy(), "品質管理": db_h, "撮影箇所": db_p}
# Sort dekigata
df = kojyo_data["出来形管理"]
sort_cols = ["編", "章", "節", "条", "枝番"]
for col in sort_cols:
    if col in df.columns:
        df[f"_s_{col}"] = df[col].apply(lambda v: float(v) if v and str(v).replace(".", "").isdigit() else float("inf"))
key_cols = [f"_s_{c}" for c in sort_cols if f"_s_{c}" in df.columns]
df = df.sort_values(key_cols).drop(columns=key_cols).reset_index(drop=True)
kojyo_data["出来形管理"] = df


PROJECTS = [
    ("北勢国道管内", "samples/input/数量総括表_北勢国道管内.pdf", "samples/output/北勢国道管内_目標output.xlsx"),
    ("麻生田",       "samples/input/数量総括表_麻生田.pdf",        "samples/output/麻生田_目標output.xlsx"),
    ("三滝川",       "samples/input/工事数量総括表_三重県_三滝川_検索可能.pdf", "samples/output/三滝川_目標output.xlsx"),
]

BASE = Path(__file__).resolve().parent.parent


def generate_output(pdf_path):
    suryo_raw = extract_suryo(str(pdf_path))
    suryo_df = suryo_raw["工種階層"]
    result_df = build_match_detail(suryo_df, db_d, db_h, db_p)

    out_d, out_h, out_p = [], [], []
    seen_d, seen_h, seen_p = set(), set(), set()
    for _, row in result_df.iterrows():
        for l in [x.strip() for x in str(row.get("出来形マッチ", "")).split("\n") if x.strip()]:
            if l not in seen_d: out_d.append(l); seen_d.add(l)
        for l in [x.strip() for x in str(row.get("品質管理マッチ", "")).split("\n") if x.strip()]:
            if l not in seen_h: out_h.append(l); seen_h.add(l)
        for l in [x.strip() for x in str(row.get("撮影箇所マッチ", "")).split("\n") if x.strip()]:
            if l not in seen_p: out_p.append(l); seen_p.add(l)

    for l in get_implicit_hinshitsu_labels(suryo_df, db_h):
        if l and l not in seen_h: out_h.append(l); seen_h.add(l)

    filtered = filter_by_row_labels(kojyo_data, out_d, out_h, out_p)
    return filtered, suryo_raw, suryo_df


def extract_hinshitsu_pairs(ws, kojyo_col, shubetsu_col):
    pairs = set()
    last_k = ""
    for r in range(5, ws.max_row + 1):
        k = norm(ws.cell(r, kojyo_col).value)
        s = norm(ws.cell(r, shubetsu_col).value)
        if k and k not in ("工種", "種別"):
            last_k = k
        if s and s not in ("工種", "種別") and last_k:
            pairs.add((last_k, s))
    return pairs


def extract_dekigata_shubetsu(ws, col=3, start=5):
    items = set()
    for r in range(start, ws.max_row + 1):
        c = norm(ws.cell(r, col).value)
        if c and c not in ("測定項目", "種別", "工種・種別", ""):
            items.add(c)
    return items


def extract_photo_kojyo(ws, start=4):
    skip = {"工種", "撮影箇所一覧表", "着手前", "完成", "工事施工中",
            "安全管理", "使用材料", "被災状況", "事故報告", "補償関係",
            "環境対策・イメージアップ等", "図面との不一致", "区分", ""}
    items = set()
    for r in range(start, ws.max_row + 1):
        c = norm(ws.cell(r, 3).value)
        if c and c not in skip:
            items.add(c)
    return items


def fuzzy_match_sets(target_set, app_set):
    covered = set()
    app_matched = set()
    for t in target_set:
        tn = norm(t) if not isinstance(t, tuple) else norm(t[0]) + norm(t[1])
        for a in app_set:
            an = norm(a) if not isinstance(a, tuple) else norm(a[0]) + norm(a[1])
            if tn == an or tn in an or an in tn:
                covered.add(t)
                app_matched.add(a)
                break
    return covered, app_matched


def compare_hinshitsu(target_wb, app_filtered):
    target_ws = target_wb["(8) 品管一覧"]
    # Target: B=工種(col2), D=種別(col4) — merged cells
    target_pairs = extract_hinshitsu_pairs(target_ws, 2, 4)
    target_pairs = {p for p in target_pairs if p[0] != "工種"}

    # App: from filtered data → reshape
    from excel_writer import _reshape_hinshitsu
    df_h = _reshape_hinshitsu(app_filtered["品質管理"])
    app_pairs = set()
    for _, row in df_h.iterrows():
        k = norm(row.get("工種", ""))
        s = norm(row.get("種別", ""))
        if k and s:
            app_pairs.add((k, s))

    covered, _ = fuzzy_match_sets(target_pairs, app_pairs)
    missing = target_pairs - covered
    # Excess: app pairs not matching any target
    _, app_matched = fuzzy_match_sets(target_pairs, app_pairs)
    excess = app_pairs - app_matched

    return target_pairs, app_pairs, covered, missing, excess


def compare_dekigata(target_wb, app_filtered):
    target_items = set()
    for sn in ["(8) 出来形一覧", "(8) 出来形一覧 (2)"]:
        if sn in target_wb.sheetnames:
            target_items |= extract_dekigata_shubetsu(target_wb[sn])

    from excel_writer import _reshape_dekigata
    df_d = _reshape_dekigata(app_filtered["出来形管理"])
    app_items = set()
    for v in df_d["種別"].dropna().unique():
        n = norm(v)
        if n and n not in ("種別", "測定項目"):
            app_items.add(n)

    covered, app_matched = fuzzy_match_sets(target_items, app_items)
    missing = target_items - covered
    excess = app_items - app_matched
    return target_items, app_items, covered, missing, excess


def compare_photo(target_wb, app_filtered):
    target_items = set()
    if "(8) 撮影箇所" in target_wb.sheetnames:
        target_items = extract_photo_kojyo(target_wb["(8) 撮影箇所"])

    from excel_writer import _reshape_photo
    df_p = _reshape_photo(app_filtered["撮影箇所"])
    # 固定セクション（着手前/完成等）はアプリ側でもスキップ
    # norm() はNFKC+空白除去のみ（括弧は残る）
    photo_skip_keywords = ["着手前", "完成", "工事施工中", "安全管理", "使用材料",
                           "被災状況", "事故報告", "補償関係", "環境対策",
                           "図面との不一致", "品質管理写真", "出来形管理写真"]
    app_items = set()
    for v in df_p["工種"].dropna().unique():
        n = norm(v)
        if not n or n in ("工種", ""):
            continue
        if any(norm(kw) in n or n in norm(kw) for kw in photo_skip_keywords):
            continue
        app_items.add(n)

    covered, app_matched = fuzzy_match_sets(target_items, app_items)
    missing = target_items - covered
    excess = app_items - app_matched
    return target_items, app_items, covered, missing, excess


# ── メイン ──
print("=" * 70)
print("差分分析レポート（現在のコード vs 目標output）")
print("=" * 70)

summary_rows = []

for name, pdf_rel, target_rel in PROJECTS:
    pdf_path = BASE / pdf_rel
    target_path = BASE / target_rel
    print(f"\n{'─' * 60}")
    print(f"■ {name}")
    print(f"{'─' * 60}")

    filtered, suryo_raw, suryo_df = generate_output(pdf_path)
    target_wb = openpyxl.load_workbook(str(target_path))

    # 品管
    t_h, a_h, cov_h, miss_h, exc_h = compare_hinshitsu(target_wb, filtered)
    print(f"\n  【品管一覧】目標:{len(t_h)} アプリ:{len(a_h)} カバー:{len(cov_h)} 欠落:{len(miss_h)} 過剰:{len(exc_h)} カバー率:{len(cov_h)/max(len(t_h),1)*100:.0f}%")
    if miss_h:
        print(f"    欠落:")
        for m in sorted(miss_h):
            print(f"      ❌ {m[0]} / {m[1]}")
    if exc_h:
        print(f"    過剰:")
        for e in sorted(exc_h):
            print(f"      ➕ {e[0]} / {e[1]}")
    summary_rows.append((name, "品管", len(t_h), len(cov_h), len(miss_h), len(exc_h)))

    # 出来形
    t_d, a_d, cov_d, miss_d, exc_d = compare_dekigata(target_wb, filtered)
    print(f"\n  【出来形一覧】目標:{len(t_d)} アプリ:{len(a_d)} カバー:{len(cov_d)} 欠落:{len(miss_d)} 過剰:{len(exc_d)} カバー率:{len(cov_d)/max(len(t_d),1)*100:.0f}%")
    if miss_d:
        print(f"    欠落:")
        for m in sorted(miss_d):
            print(f"      ❌ {m}")
    if exc_d:
        print(f"    過剰:")
        for e in sorted(exc_d):
            print(f"      ➕ {e}")
    summary_rows.append((name, "出来形", len(t_d), len(cov_d), len(miss_d), len(exc_d)))

    # 撮影箇所
    t_p, a_p, cov_p, miss_p, exc_p = compare_photo(target_wb, filtered)
    print(f"\n  【撮影箇所】目標:{len(t_p)} アプリ:{len(a_p)} カバー:{len(cov_p)} 欠落:{len(miss_p)} 過剰:{len(exc_p)} カバー率:{len(cov_p)/max(len(t_p),1)*100:.0f}%")
    if miss_p:
        print(f"    欠落:")
        for m in sorted(miss_p):
            print(f"      ❌ {m}")
    if exc_p:
        print(f"    過剰(先頭15件):")
        for e in sorted(exc_p)[:15]:
            print(f"      ➕ {e}")
        if len(exc_p) > 15:
            print(f"      ... 他{len(exc_p)-15}件")
    summary_rows.append((name, "撮影箇所", len(t_p), len(cov_p), len(miss_p), len(exc_p)))

# サマリー
print(f"\n{'=' * 70}")
print("サマリー")
print(f"{'=' * 70}")
print(f"{'工事':<12} {'シート':<8} {'目標':>4} {'カバー':>6} {'欠落':>4} {'過剰':>4} {'カバー率':>8}")
print("-" * 50)
total_target = total_cover = total_miss = total_exc = 0
for name, sheet, tgt, cov, mis, exc in summary_rows:
    rate = f"{cov/max(tgt,1)*100:.0f}%"
    print(f"{name:<12} {sheet:<8} {tgt:>4} {cov:>6} {mis:>4} {exc:>4} {rate:>8}")
    total_target += tgt
    total_cover += cov
    total_miss += mis
    total_exc += exc
print("-" * 50)
print(f"{'合計':<12} {'':8} {total_target:>4} {total_cover:>6} {total_miss:>4} {total_exc:>4} {total_cover/max(total_target,1)*100:.0f}%")
