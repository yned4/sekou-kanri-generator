"""
excel_writer.py
抽出済みDataFrameを施工管理計画Excelファイルに書き出す。
手動追加行（custom_rows）対応済み。

使い方（単体テスト）:
    python excel_writer.py  # extractor.py と組み合わせて動作確認
"""

import io
import re
from typing import Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===== シート名定義 =====
SHEET_HINSHITSU  = "(8) 品管一覧"
SHEET_DEKIGATA   = "(8) 出来形一覧"
SHEET_DEKIGATA2  = "(8) 出来形一覧 (2)"
SHEET_PHOTO      = "(8) 撮影箇所"
SHEET_ALERT      = "DB未登録"

# シート固有タイトル（行1 B列に表示）
_SHEET_TITLE = {
    SHEET_HINSHITSU: "品質管理基準及び規格値",
    SHEET_DEKIGATA:  "出来形管理基準及び規格値",
    SHEET_DEKIGATA2: "出来形管理基準及び規格値",
    SHEET_PHOTO:     "撮影箇所一覧表",
}

# 出来形1シートあたりの行上限（超えたら2枚に分割）
# 目標outputでは約39行で分割されているため、それに近い値を設定
DEKIGATA_SPLIT_THRESHOLD = 35

# ===== スタイル定数 =====
HEADER_BG_COLOR  = "D1D1D1"   # グレースケール82%
HEADER_FONT_COLOR = "000000"  # 黒文字
SECTION_BG_COLOR = "F0F0F0"   # 薄いグレー（セクション区切り行）
FONT_NAME        = "ＭＳ 明朝" # 出力フォント（全角）
MAX_COL_WIDTH    = 30          # 列幅の上限（文字数相当）
MIN_COL_WIDTH    = 8           # 列幅の下限
A_COL_WIDTH      = 4           # A列余白の幅（文字数相当）
TATE_COL_WIDTH   = 4           # 縦書き列の幅（文字数相当）


def _make_header_style() -> tuple:
    """ヘッダー行用スタイルを返す。"""
    font  = Font(name=FONT_NAME, bold=False, color=HEADER_FONT_COLOR, size=8)
    fill  = PatternFill("solid", fgColor=HEADER_BG_COLOR)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return font, fill, align


def _make_data_align(wrap: bool = True) -> Alignment:
    return Alignment(vertical="center", wrap_text=wrap)


def _make_thin_border() -> Border:
    thin = Side(style="thin", color="AAAAAA")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _merge_col_cells(ws, col_idx: int, data_start_row: int, data_end_row: int,
                     tate: bool = False, break_rows: set = None) -> None:
    """
    指定列の連続する同値セルを縦結合する。
    結合後は上端セルに値・スタイルを保持し、中央揃えにする。
    tate=True の場合は縦書き（textRotation=255）を適用する。
    break_rows: 親列の値が変わる行番号のセット。この行では値が同じでも結合を切る。
    """
    if data_end_row < data_start_row:
        return
    thin = Side(style="thin", color="AAAAAA")
    merged_align = (
        Alignment(textRotation=255, horizontal="center", vertical="center")
        if tate else
        Alignment(horizontal="left", vertical="center", wrap_text=True)
    )

    row = data_start_row
    while row <= data_end_row:
        cell_val = ws.cell(row=row, column=col_idx).value
        # 同じ値が続く範囲を探す（親列の境界では切る）
        end_row = row
        while (end_row + 1 <= data_end_row
               and ws.cell(row=end_row + 1, column=col_idx).value == cell_val
               and (break_rows is None or end_row + 1 not in break_rows)):
            end_row += 1

        if end_row > row:
            # 複数行 → セル結合
            ws.merge_cells(
                start_row=row, start_column=col_idx,
                end_row=end_row, end_column=col_idx,
            )
            top_cell = ws.cell(row=row, column=col_idx)
            top_cell.alignment = merged_align
            # 結合セルの外枠ボーダーを設定
            top_cell.border = Border(
                left=thin, right=thin, top=thin, bottom=thin
            )

        row = end_row + 1


def _write_sheet(
    ws,
    df: pd.DataFrame,
    section_col: Optional[str] = None,
    title: str = "",
    merge_cols: Optional[list] = None,
    tate_cols: Optional[list] = None,
) -> None:
    """
    DataFrameを1シートに書き込む。

    シート構造（目標outputに合わせた形式）:
      行1:  A1=ページ番号(空欄)  B1=タイトル
      行2:  空行
      行3:  空行
      行4:  ヘッダー行（A列=余白列、B列以降=列名）
      行5:  ヘッダー続き行（高さ確保のため空セルにヘッダースタイルを適用）
      行6〜: データ行（A列=余白、B列以降=データ）

    section_col: この列の値が変わるタイミングでセクション区切り行の背景色を変える。
    title: 行1 B列に表示するシート固有タイトル文字列。
    merge_cols: 連続する同値セルを縦結合する列名リスト。
    tate_cols: 縦書き（textRotation=255）を適用する列名リスト（ヘッダー除く）。
    """
    header_font, header_fill, header_align = _make_header_style()
    data_align  = _make_data_align()
    tate_align  = Alignment(textRotation=255, horizontal="center", vertical="center")
    thin_border = _make_thin_border()
    tate_col_set = set(tate_cols or [])

    n_data_cols = len(df.columns)

    # ── 行1: タイトル行 ─────────────────────────────────────
    # A1 = ページ番号プレースホルダー（空欄）
    ws.cell(row=1, column=1, value="")
    # B1 = シート固有タイトル（データ列の末尾まで結合）
    title_cell = ws.cell(row=1, column=2, value=title)
    title_cell.font      = Font(name=FONT_NAME, bold=False, size=8)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    if n_data_cols > 1:
        ws.merge_cells(
            start_row=1, start_column=2,
            end_row=1,   end_column=n_data_cols + 1,
        )
    ws.row_dimensions[1].height = 18

    # ── 行2〜3: 空行 ────────────────────────────────────────
    ws.row_dimensions[2].height = 10
    ws.row_dimensions[3].height = 10

    # ── 行4: ヘッダー行 ─────────────────────────────────────
    # A4 = 余白列（フィル・ボーダーなし）
    a4 = ws.cell(row=4, column=1, value="")
    a4.fill = PatternFill(fill_type=None)  # 明示的にフィルなし

    for col_idx, col_name in enumerate(df.columns, start=2):  # B列=2
        cell = ws.cell(row=4, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border

    ws.row_dimensions[4].height = 30

    # ── 行5: ヘッダー続き行（高さ確保用・スタイルのみ） ──────
    # A5 = 余白列（フィル・ボーダーなし）
    a5 = ws.cell(row=5, column=1, value="")
    a5.fill = PatternFill(fill_type=None)  # 明示的にフィルなし
    for col_idx in range(2, n_data_cols + 2):  # B列以降のみ
        cell = ws.cell(row=5, column=col_idx, value="")
        cell.fill   = header_fill
        cell.border = thin_border
    ws.row_dimensions[5].height = 10

    ws.freeze_panes = "B6"

    # ── 行6〜: データ行 ──────────────────────────────────────
    section_fill = PatternFill("solid", fgColor=SECTION_BG_COLOR)
    prev_section = None

    for row_idx, (_, row) in enumerate(df.iterrows(), start=6):
        current_section = row.get(section_col) if section_col else None
        use_section_fill = (
            section_col is not None
            and current_section != prev_section
            and current_section is not None
        )

        # A列（余白列）
        a_cell = ws.cell(row=row_idx, column=1, value="")
        a_cell.font   = Font(name=FONT_NAME, size=8)
        a_cell.border = thin_border

        for col_idx, (col_name, value) in enumerate(zip(df.columns, row), start=2):  # B列=2
            if pd.isna(value) if not isinstance(value, str) else False:
                value = ""
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value != "" else "")
            cell.font      = Font(name=FONT_NAME, size=8)
            cell.alignment = tate_align if col_name in tate_col_set else data_align
            cell.border    = thin_border
            if use_section_fill:
                cell.fill = section_fill

        if use_section_fill:
            prev_section = current_section

    # ── 列幅の自動調整 ──────────────────────────────────────
    ws.column_dimensions["A"].width = A_COL_WIDTH

    for col_idx, col_name in enumerate(df.columns, start=2):  # B列=2
        col_letter = get_column_letter(col_idx)
        if col_name in tate_col_set:
            # 縦書き列は固定幅（内容の長さで広げない）
            ws.column_dimensions[col_letter].width = TATE_COL_WIDTH
        else:
            max_len = len(str(col_name))
            for value in df.iloc[:, col_idx - 2]:  # df列は0始まり
                if pd.isna(value) if not isinstance(value, str) else False:
                    continue
                cell_max = max((len(line) for line in str(value).split("\n")), default=0)
                max_len = max(max_len, cell_max)
            width = min(max(max_len + 2, MIN_COL_WIDTH), MAX_COL_WIDTH)
            ws.column_dimensions[col_letter].width = width

    # ── 指定列の縦セル結合 ────────────────────────────────────
    # 親列の値が変わる行で子列の結合を切る（例: 工種が変われば種別も分離）
    if merge_cols and len(df) > 0:
        data_end_row = 5 + len(df)  # ヘッダー5行 + データ行数
        col_name_to_idx = {name: idx for idx, name in enumerate(df.columns, start=2)}
        accumulated_breaks: set[int] = set()
        for col_name in merge_cols:
            if col_name not in col_name_to_idx:
                continue
            _merge_col_cells(ws, col_name_to_idx[col_name], 6, data_end_row,
                             tate=col_name in tate_col_set,
                             break_rows=accumulated_breaks or None)
            # この列の値が変わる行を後続列の境界として蓄積
            if col_name in df.columns:
                col_vals = df[col_name].tolist()
                for i in range(1, len(col_vals)):
                    if col_vals[i] != col_vals[i - 1]:
                        accumulated_breaks.add(6 + i)  # データ行は row 6 開始



def _compute_shauchi(規格値_val: str) -> str:
    """
    社内規格値を算出する。

    ・数値が含まれる場合: 各数値に0.8を乗じる（小数桁数は元に合わせる）。
    ・数値が含まれない場合: 規格値をそのまま転記。
    """
    s = str(規格値_val or "").strip()
    if not s or not re.search(r'\d', s):
        return s

    def _scale(m: re.Match) -> str:
        orig = m.group()
        val = float(orig)
        scaled = val * 0.8
        if '.' in orig:
            decimals = len(orig.split('.')[1])
            return f"{round(scaled, decimals):.{decimals}f}"
        else:
            # 元が整数でも結果が小数になる場合は小数1桁で表示
            if scaled == int(scaled):
                return str(int(scaled))
            return f"{scaled:.1f}"

    return re.sub(r'\d+(?:\.\d+)?', _scale, s)


# 品質管理工種の表示名マッピング
# matching_rules.json から読み込む（対応表編集UIまたはJSON直接編集で変更可能）
from matching_rules import get_hinshitsu_kojyo_display as _get_display_map


def _reshape_hinshitsu(df: pd.DataFrame) -> pd.DataFrame:
    """
    品質管理DataFrameを出力列形式に変換する。

    工種・種別は国交省DB品質管理シートの値をそのまま使用する。
    （出来形と異なり、数量総括表の工種へのマッピングは行わない）

    削除: 試験区分、試験成績表等による確認
    追加: 社内規格値（空欄）
    リネーム: 試験時期・頻度→試験基準、摘要→備考
    """
    out = pd.DataFrame(index=df.index)
    # 国交省DB品質管理シートの工種名を使用（表示名マッピングで変換）
    display_map = _get_display_map()
    out["工種"]       = df["工種"].fillna("").apply(
        lambda x: display_map.get(x, x)
    )
    out["種別"]       = df["種別"].fillna("")
    out["試験項目"]   = df["試験項目"].fillna("")
    out["試験方法"]   = df["試験方法"].fillna("")
    out["規格値"]     = df["規格値"].fillna("")
    out["社内規格値"] = df["規格値"].apply(_compute_shauchi)
    out["試験基準"]   = df["試験時期・頻度"].fillna("") if "試験時期・頻度" in df.columns else ""
    out["備考"]       = df["摘要"].fillna("") if "摘要" in df.columns else ""
    # 表示名マッピング後に重複行（工種+種別+試験項目が同一）を除去
    out = out.drop_duplicates(subset=["工種", "種別", "試験項目"], keep="first")
    return out.reset_index(drop=True)


def _reshape_dekigata(df: pd.DataFrame, kojyo_to_suryo: dict = None) -> pd.DataFrame:
    """
    出来形管理DataFrameを出力列形式に変換する。
    削除: 編、章、節、条、枝番、規格値_個々、測定基準
    追加: 工種（数量総括表由来の大分類）、社内規格値（空欄）
    リネーム: 工種→種別、摘要→備考

    kojyo_to_suryo: {国交省工種名: 数量総括表工種名} の対応辞書
    """
    out = pd.DataFrame()
    if kojyo_to_suryo:
        out["工種"] = df["工種"].map(lambda x: kojyo_to_suryo.get(str(x), "") if x else "")
    else:
        out["工種"] = ""
    out["種別"]        = df.get("工種", "")
    out["測定項目"]    = df.get("測定項目", "")
    out["規格値_条件"] = df.get("規格値_条件", "")
    out["規格値"]      = df.get("規格値", "")
    out["社内規格値"]  = df["規格値"].apply(_compute_shauchi)
    out["測定箇所"]    = df.get("測定基準", "")
    out["備考"]        = df.get("摘要", "")
    return out


_SECTION_TO_KUBUN = {
    "出来形管理": "出来形管理写真",
    "品質管理":   "品質管理写真",
}

# 区分ごとの提出頻度ルール（目標outputより）
_TEISHUTSU_RULES: list[tuple[str, str]] = [
    ("着手前",       "着手前1枚"),
    ("完成",         "施工完了後1枚"),
    ("工事施工中",   "不要"),
    ("安全管理",     "全景1枚"),
    ("使用材料",     "不要"),
    ("品質管理写真", "不要"),
    ("出来形管理写真", "不要"),
]


def _get_teishutsu(kubun) -> str:
    """区分文字列からルールベースで提出頻度を返す。"""
    kubun = str(kubun or "")
    for key, val in _TEISHUTSU_RULES:
        if key in kubun:
            return val
    return ""


def _parse_photo_frequency(freq_str: str) -> tuple:
    """
    撮影頻度文字列から (撮影時期, 撮影頻度) を分離して返す。

    国交省基準PDFの形式:
      全体セクション:    "着手前１回 〔着手前〕"      → ("着手前", "着手前１回")
      品質/出来形管理:   "各種路盤毎に１回 [試験実施中]" → ("試験実施中", "各種路盤毎に１回")

    〔〕または [] の最初のブロックだけを撮影時期として抽出し、
    残りの文字列（ただし書き含む）を撮影頻度とする。
    括弧が見つからない場合は全体を撮影頻度とし、撮影時期は空文字。
    """
    s = str(freq_str or "").strip()
    m = re.search(r'[〔\[](.+?)[〕\]]', s)
    if m:
        timing = m.group(1).strip()
        # 〔〕/[] ブロックとその前のスペースを除去し、後続テキストは1スペースで繋ぐ
        freq = re.sub(r'\s*[〔\[].+?[〕\]]\s*', ' ', s, count=1).strip()
        return timing, freq
    return "", s


def _reshape_photo(df: pd.DataFrame) -> pd.DataFrame:
    """
    撮影箇所DataFrameを出力列形式に変換する。

    列構成（目標outputに合わせ7列）:
      区分 / 工種 / 撮影項目 / 撮影時期 / 撮影頻度 / 提出頻度 / 摘要

    区分:
      全体セクション  → 区分列の値をそのまま使用
      品質/出来形セクション → セクション名を「品質管理写真」「出来形管理写真」に変換

    工種:
      全体セクション  → sub区分列の値を使用（「着手前」「工事施工中」等）
      品質/出来形セクション → 工種列の値を使用

    撮影時期・撮影頻度:
      DBの「撮影頻度」列から 〔〕/[] を手がかりに自動分離する。
      提出頻度はDBに情報がないため空欄とする。
    """
    records = []

    for _, row in df.iterrows():
        sec = str(row.get("セクション") or "")

        # 全体セクションの除外条件
        if sec == "全体":
            # ① sub区分・撮影項目・撮影頻度がすべて空 → PDF上のカテゴリ見出し行
            if (not (row.get("sub区分") or "")
                    and not (row.get("撮影項目") or "")
                    and not (row.get("撮影頻度") or "")):
                continue
            # ② 区分が出来形管理・品質管理 → 専用セクションで詳細に出力済みのため重複除外
            raw_k = str(row.get("区分") or "")
            if raw_k in ("出来形管理", "品質管理"):
                continue

        # 区分
        raw_kubun = row.get("区分") or ""
        kubun = raw_kubun if raw_kubun else _SECTION_TO_KUBUN.get(sec, sec)

        # 工種
        raw_kojyo = row.get("工種") or ""
        kojyo = raw_kojyo if raw_kojyo else (row.get("sub区分") or "")

        # 撮影時期・撮影頻度を分離
        timing, freq = _parse_photo_frequency(row.get("撮影頻度") or "")

        # 「着手前・完成」は着手前・完成の2行に分離（#10）
        kubun_list = ["着手前", "完成"] if kubun == "着手前・完成" else [kubun]

        for kb in kubun_list:
            records.append({
                "区分":     kb,
                "工種":     kojyo,
                "撮影項目": row.get("撮影項目") or "",
                "撮影時期": timing,
                "撮影頻度": freq,
                "提出頻度": _get_teishutsu(kb),  # ルールベースで補完（#11）
                "摘要":     row.get("摘要") or "",
            })

    out = pd.DataFrame(
        records,
        columns=["区分", "工種", "撮影項目", "撮影時期", "撮影頻度", "提出頻度", "摘要"],
    )
    # 「着手前・完成」分離により生じる重複行を除去
    out = out.drop_duplicates().reset_index(drop=True)
    return out


def _inject_custom_rows(df: pd.DataFrame, custom_rows: list, sheet_key: str) -> pd.DataFrame:
    """
    カスタム行をDataFrameの指定位置に挿入する。

    after_kojyo == "末尾" → DataFrame末尾に追加
    after_kojyo == "<工種名>" → その工種の最終行の直後に挿入（見つからなければ末尾）
    """
    target = [r for r in custom_rows if r.get("sheet") == sheet_key]
    if not target or df.empty:
        for row in target:
            new_df = pd.DataFrame([{c: row.get("fields", {}).get(c, "") for c in df.columns}])
            df = pd.concat([df, new_df], ignore_index=True)
        return df

    for row in target:
        fields = row.get("fields", {})
        new_row = {c: fields.get(c, "") for c in df.columns}
        new_df = pd.DataFrame([new_row])

        after = row.get("after_kojyo", "末尾")
        if after != "末尾" and "工種" in df.columns:
            mask = df["工種"].astype(str) == after
            if mask.any():
                insert_at = int(df[mask].index[-1]) + 1
                # reset_index後のDataFrameでのloc位置
                pos = df.index.get_loc(df.index[insert_at - 1]) + 1 if insert_at <= len(df) else len(df)
                df = pd.concat(
                    [df.iloc[:pos], new_df, df.iloc[pos:]],
                    ignore_index=True,
                )
                continue
        df = pd.concat([df, new_df], ignore_index=True)

    return df


def _write_alert_sheet(ws, unmatched_items: list,
                       skip_keywords: Optional[list] = None) -> None:
    """DB未登録の工種をアラートシートに書き出す（対象外/要確認の2段階）。"""
    import matching_rules as mr
    if skip_keywords is None:
        skip_keywords = mr.get_skip_alert_keywords()

    # 対象外 / 要確認 に分類
    warn_items, skip_items = [], []
    for item in unmatched_items:
        joined = "".join(str(item.get(c, "")) for c in ["工種", "種別", "細別", "名称"])
        if any(kw in joined for kw in skip_keywords):
            skip_items.append(item)
        else:
            warn_items.append(item)

    warn_fill = PatternFill("solid", fgColor="FFF3CD")
    warn_font = Font(name=FONT_NAME, bold=True, size=9, color="856404")
    note_font = Font(name=FONT_NAME, size=8, color="856404")
    skip_fill = PatternFill("solid", fgColor="F0F0F0")
    skip_font = Font(name=FONT_NAME, bold=True, size=9, color="666666")
    skip_note_font = Font(name=FONT_NAME, size=8, color="666666")
    hdr_font  = Font(name=FONT_NAME, bold=False, size=8, color="000000")
    hdr_fill  = PatternFill("solid", fgColor=HEADER_BG_COLOR)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align = _make_data_align()
    border     = _make_thin_border()
    cols = ["工種", "種別", "細別", "名称"]

    row = 1

    # ── 要確認セクション ──
    if warn_items:
        c = ws.cell(row=row, column=1,
                    value=f"DB未登録 — 要確認（{len(warn_items)} 件）")
        c.font, c.fill = warn_font, warn_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 28
        row += 1

        c = ws.cell(row=row, column=1,
                    value="以下の工種は国交省DBに該当がないため自動出力されません。"
                          "手動行の追加、または対応表の更新で対応してください。")
        c.font = note_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 20
        row += 1

        ws.row_dimensions[row].height = 8
        row += 1

        for ci, col_name in enumerate(cols, start=1):
            c = ws.cell(row=row, column=ci, value=col_name)
            c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, border
        ws.row_dimensions[row].height = 24
        row += 1

        for item in warn_items:
            for ci, col_name in enumerate(cols, start=1):
                c = ws.cell(row=row, column=ci, value=item.get(col_name, ""))
                c.font = Font(name=FONT_NAME, size=8)
                c.alignment = data_align
                c.border = border
            row += 1

        row += 1  # 空行

    # ── 対象外セクション ──
    if skip_items:
        c = ws.cell(row=row, column=1,
                    value=f"施工管理基準の対象外（{len(skip_items)} 件）")
        c.font, c.fill = skip_font, skip_fill
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 28
        row += 1

        c = ws.cell(row=row, column=1,
                    value="以下の工種は仮設・撤去等のため施工管理基準の対象外です。対応は不要です。")
        c.font = skip_note_font
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.row_dimensions[row].height = 20
        row += 1

        ws.row_dimensions[row].height = 8
        row += 1

        for ci, col_name in enumerate(cols, start=1):
            c = ws.cell(row=row, column=ci, value=col_name)
            c.font, c.fill, c.alignment, c.border = hdr_font, hdr_fill, hdr_align, border
        ws.row_dimensions[row].height = 24
        row += 1

        for item in skip_items:
            for ci, col_name in enumerate(cols, start=1):
                c = ws.cell(row=row, column=ci, value=item.get(col_name, ""))
                c.font = Font(name=FONT_NAME, size=8, color="666666")
                c.alignment = data_align
                c.border = border
            row += 1

    # 列幅
    for ci, w in enumerate([24, 20, 20, 28], start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def _build_wb(df_d: pd.DataFrame, df_h: pd.DataFrame, df_p: pd.DataFrame,
              unmatched_items: list = None) -> "Workbook":
    """DataFrameからWorkbookを構築する（内部ヘルパー）。"""
    wb = Workbook()
    wb.remove(wb.active)

    ws_h = wb.create_sheet(SHEET_HINSHITSU)
    _write_sheet(ws_h, df_h, title=_SHEET_TITLE[SHEET_HINSHITSU],
                 merge_cols=["工種", "種別"], tate_cols=["工種", "種別"])

    if len(df_d) > DEKIGATA_SPLIT_THRESHOLD:
        df_d1 = df_d.iloc[:DEKIGATA_SPLIT_THRESHOLD].reset_index(drop=True)
        df_d2 = df_d.iloc[DEKIGATA_SPLIT_THRESHOLD:].reset_index(drop=True)
        ws_d1 = wb.create_sheet(SHEET_DEKIGATA)
        _write_sheet(ws_d1, df_d1, title=_SHEET_TITLE[SHEET_DEKIGATA],
                     merge_cols=["工種", "種別"], tate_cols=["工種"])
        ws_d2 = wb.create_sheet(SHEET_DEKIGATA2)
        _write_sheet(ws_d2, df_d2, title=_SHEET_TITLE[SHEET_DEKIGATA2],
                     merge_cols=["工種", "種別"], tate_cols=["工種"])
    else:
        ws_d = wb.create_sheet(SHEET_DEKIGATA)
        _write_sheet(ws_d, df_d, title=_SHEET_TITLE[SHEET_DEKIGATA],
                     merge_cols=["工種", "種別"], tate_cols=["工種"])

    # 撮影箇所: 区分の正規順でグループ化し、同一区分内は工種の初出順でまとめる
    if not df_p.empty:
        _KUBUN_ORDER = [
            "着手前", "完成", "施工状況", "安全管理", "使用材料",
            "品質管理写真", "出来形管理写真", "災害", "事故", "その他",
        ]
        kubun_rank = {v: i for i, v in enumerate(_KUBUN_ORDER)}
        df_p = df_p.copy()
        df_p["_kubun_rank"] = df_p["区分"].map(lambda x: kubun_rank.get(x, 999))
        # 同一区分内で工種の初出順を保持しつつ同一工種行をまとめる
        kojyo_first = {}
        kojyo_ranks = []
        for _, r in df_p.iterrows():
            key = (r["_kubun_rank"], r["工種"])
            if key not in kojyo_first:
                kojyo_first[key] = len(kojyo_first)
            kojyo_ranks.append(kojyo_first[key])
        df_p["_kojyo_rank"] = kojyo_ranks
        df_p = df_p.sort_values(["_kubun_rank", "_kojyo_rank"], kind="stable")
        df_p = df_p.drop(columns=["_kubun_rank", "_kojyo_rank"]).reset_index(drop=True)

    ws_p = wb.create_sheet(SHEET_PHOTO)
    _write_sheet(ws_p, df_p, title=_SHEET_TITLE[SHEET_PHOTO],
                 merge_cols=list(df_p.columns))

    # DB未登録工種のアラートシート
    if unmatched_items:
        ws_alert = wb.create_sheet(SHEET_ALERT)
        _write_alert_sheet(ws_alert, unmatched_items)
        ws_alert.sheet_properties.tabColor = "FFC107"  # タブ色: 黄色

    return wb


def write_excel(
    data: dict,
    output_path: Optional[str] = None,
    工事名: str = "",
    dekigata_kojyo_map: dict = None,
    custom_rows: Optional[list] = None,
    return_dfs: bool = False,
    unmatched_items: Optional[list] = None,
):
    """
    抽出データをExcelに書き出す。

    Args:
        data:               filter_by_kojyo() または extract_from_pdf() の戻り値
        output_path:        保存先パス。None の場合はバイト列を返す（Streamlit用）。
        工事名:             数量総括表から取得した工事名。シート先頭行に表示する。
        dekigata_kojyo_map: {国交省工種名: 数量総括表工種名} の対応辞書。
        custom_rows:        手動追加行リスト。
        return_dfs:         True の場合 (bytes, dfs_dict) を返す。
        unmatched_items:    DB未登録の数量総括表工種リスト（アラートシート用）。

    Returns:
        return_dfs=False: Excel バイト列（output_pathがNoneの場合）
        return_dfs=True:  (bytes, {"出来形一覧": df_d, "品管一覧": df_h, "撮影箇所": df_p})
    """
    df_h = _reshape_hinshitsu(data["品質管理"])
    df_d = _reshape_dekigata(data["出来形管理"], kojyo_to_suryo=dekigata_kojyo_map)
    df_p = _reshape_photo(data["撮影箇所"])

    if custom_rows:
        df_h = _inject_custom_rows(df_h, custom_rows, "品管一覧")
        df_d = _inject_custom_rows(df_d, custom_rows, "出来形一覧")
        df_p = _inject_custom_rows(df_p, custom_rows, "撮影箇所")

    wb = _build_wb(df_d, df_h, df_p, unmatched_items=unmatched_items)

    if output_path:
        wb.save(output_path)
        result = None
    else:
        buf = io.BytesIO()
        wb.save(buf)
        result = buf.getvalue()

    if return_dfs:
        return result, {"出来形一覧": df_d.copy(), "品管一覧": df_h.copy(), "撮影箇所": df_p.copy()}
    return result


def write_excel_from_dfs(
    df_d: pd.DataFrame,
    df_h: pd.DataFrame,
    df_p: pd.DataFrame,
    output_path: Optional[str] = None,
) -> bytes:
    """
    編集済みDataFrameから直接Excelを生成する（Excel編集ページ用）。

    Args:
        df_d: 出来形一覧 DataFrame
        df_h: 品管一覧 DataFrame
        df_p: 撮影箇所 DataFrame
        output_path: 保存先パス。None の場合はバイト列を返す。
    """
    wb = _build_wb(df_d, df_h, df_p)

    if output_path:
        wb.save(output_path)
        return None
    else:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


# ---------------------------------------------------------------------------
# 単体動作確認
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    from extractor import extract_from_pdf

    施工管理基準 = "/Users/tomoki/Downloads/OneDrive_1_2026-5-18/土木工事施工管理基準及び規格値（案）.pdf"
    写真管理基準 = "/Users/tomoki/Downloads/OneDrive_1_2026-5-18/写真管理基準.pdf"
    output      = "output_test.xlsx"

    print("抽出中...")
    data = extract_from_pdf(施工管理基準, 写真管理基準)

    print("Excel生成中...")
    write_excel(data, output_path=output)
    print(f"保存完了: {output}")
