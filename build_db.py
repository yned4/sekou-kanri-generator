"""
build_db.py
国交省PDFからExcelデータベースを構築する。

国交省基準PDFが改定された際（2〜3年に1回）だけ実行する。
通常のアプリ利用時には不要。

使い方:
    python build_db.py <施工管理基準.pdf> <写真管理基準.pdf> [--version "令和7年3月版"]

引数省略時はスクリプト内のデフォルトパスを使用。
"""

import argparse
import sys
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

from extractor import extract_from_pdf

# ===== 設定 =====
DB_PATH = Path(__file__).parent / "database" / "kojyo_kijun.xlsx"

# データシート名（app.py と共有する定数）
SHEET_DEKIGATA  = "出来形管理"
SHEET_HINSHITSU = "品質管理"
SHEET_PHOTO     = "撮影箇所"
SHEET_VERSION   = "バージョン情報"

# デフォルトPDFパス（ローカル確認用）
DEFAULT_施工管理基準 = (
    "/Users/tomoki/Downloads/OneDrive_1_2026-5-18"
    "/土木工事施工管理基準及び規格値（案）.pdf"
)
DEFAULT_写真管理基準 = (
    "/Users/tomoki/Downloads/OneDrive_1_2026-5-18"
    "/写真管理基準.pdf"
)


def _write_header(ws, columns: list) -> None:
    """ヘッダー行を書き込む（太字・背景色）。"""
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF", size=10)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill      = fill
        cell.font      = font
        cell.alignment = align
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "A2"


def _write_df(ws, df: pd.DataFrame) -> None:
    """DataFrameをシートに書き込む（ヘッダー行 + データ行）。"""
    _write_header(ws, list(df.columns))
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            v = "" if (value is None or (isinstance(value, float) and pd.isna(value))) else str(value)
            ws.cell(row=row_idx, column=col_idx, value=v)


def build(
    path_施工管理基準: str,
    path_写真管理基準: str,
    version: str,
    note: str = "",
) -> None:
    """
    2つのPDFからデータを抽出してExcelデータベースを作成・更新する。

    既存のDBファイルが存在する場合は上書き（最新版で置き換え）する。
    旧バージョンとの差分が必要な場合は手動でバックアップしてから実行すること。
    """
    print(f"[1/3] PDFを抽出中...")
    data = extract_from_pdf(path_施工管理基準, path_写真管理基準)

    print(f"[2/3] Excelデータベースを書き込み中: {DB_PATH}")
    DB_PATH.parent.mkdir(exist_ok=True)

    with pd.ExcelWriter(str(DB_PATH), engine="openpyxl") as writer:
        data["出来形管理"].to_excel(writer, sheet_name=SHEET_DEKIGATA,  index=False)
        data["品質管理"].to_excel(  writer, sheet_name=SHEET_HINSHITSU, index=False)
        data["撮影箇所"].to_excel(  writer, sheet_name=SHEET_PHOTO,     index=False)

        # バージョン情報シート
        df_ver = pd.DataFrame([{
            "バージョン":       version,
            "施工管理基準PDF":  Path(path_施工管理基準).name,
            "写真管理基準PDF":  Path(path_写真管理基準).name,
            "作成日時":         datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "メモ":             note,
        }])
        df_ver.to_excel(writer, sheet_name=SHEET_VERSION, index=False)

    # ヘッダースタイルを適用
    wb = load_workbook(str(DB_PATH))
    for sheet_name, df in [
        (SHEET_DEKIGATA,  data["出来形管理"]),
        (SHEET_HINSHITSU, data["品質管理"]),
        (SHEET_PHOTO,     data["撮影箇所"]),
    ]:
        _write_header(wb[sheet_name], list(df.columns))
    _write_header(wb[SHEET_VERSION], ["バージョン", "施工管理基準PDF", "写真管理基準PDF", "作成日時", "メモ"])
    wb.save(str(DB_PATH))

    print(f"[3/3] 完了!")
    print(f"      出来形管理: {len(data['出来形管理'])}行")
    print(f"      品質管理:   {len(data['品質管理'])}行")
    print(f"      撮影箇所:   {len(data['撮影箇所'])}行")
    print(f"      保存先:     {DB_PATH}")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="国交省基準PDFからExcelDBを構築する")
    parser.add_argument("施工管理基準", nargs="?", default=DEFAULT_施工管理基準)
    parser.add_argument("写真管理基準", nargs="?", default=DEFAULT_写真管理基準)
    parser.add_argument("--version", default="令和7年3月版", help="バージョン名（例: 令和7年3月版）")
    parser.add_argument("--note",    default="",           help="改定メモ（任意）")
    args = parser.parse_args()

    build(
        path_施工管理基準=args.施工管理基準,
        path_写真管理基準=args.写真管理基準,
        version=args.version,
        note=args.note,
    )
