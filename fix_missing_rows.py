"""
fix_missing_rows.py
出来形管理DB (kojyo_kijun.xlsx) の規格値マス欠落を動的に検出・修正するスクリプト。

処理:
1. 各ケースの工種ブロックをDB内で動的に検索
2. 既存行と比較して欠落行のみ挿入
3. ヘッダースタイル・バージョン情報を更新
"""

import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from copy import copy

DB_PATH = "database/kojyo_kijun.xlsx"
SHEET_NAME = "出来形管理"
VERSION_SHEET = "バージョン情報"


def define_cases():
    """全131ケースの欠落行定義を返す。

    各ケースは dict:
      'jou': 条 (str), 'eda': 枝番 (str), 'koushu': 工種名 (部分一致検索用),
      'rows': [{'col7': 測定項目, 'col8': 規格値_条件, 'col9': 規格値, 'col10': 規格値_個々,
                'col6': 工種名上書き(optional)}]
      'section_hint': 節の部分文字列 (同名工種の区別用, optional)
      'pdf_page': PDF頁 (参考)
    """
    cases = []

    # Case 1: 掘削工 (河川) 条2枝番1
    cases.append({'jou': '２', 'eda': '１', 'koushu': '掘削工',
                  'section_hint': '河 川', 'pdf_page': 2,
                  'rows': [
                      {'col7': '法長ℓ', 'col8': 'ℓ≧5m', 'col9': '法長の－4％', 'col10': None},
                  ]})

    # Case 2: 盛土工 条3枝番1
    cases.append({'jou': '３', 'eda': '１', 'koushu': '盛土工',
                  'section_hint': '河 川', 'pdf_page': 4,
                  'rows': [
                      {'col7': '幅ｗ', 'col8': None, 'col9': '－100', 'col10': None},
                  ]})

    # Case 3: 盛土工（面管理の場合） 条3枝番2
    cases.append({'jou': '３', 'eda': '２', 'koushu': '盛土工（面管理の場合）',
                  'section_hint': '河 川', 'pdf_page': 5,
                  'rows': [
                      {'col7': '法面 4割≧勾配 (小段含む)', 'col8': '標高較差', 'col9': '－60', 'col10': '-170'},
                  ]})

    # Case 4: 堤防天端工 条6
    cases.append({'jou': '６', 'eda': '', 'koushu': '堤防天端工',
                  'section_hint': '河 川', 'pdf_page': 6,
                  'rows': [
                      {'col7': '幅ｗ', 'col8': None, 'col9': '－100', 'col10': None},
                  ]})

    # Case 5: 掘削工 (道路) 条2枝番1
    cases.append({'jou': '２', 'eda': '１', 'koushu': '掘削工',
                  'section_hint': '道 路', 'pdf_page': 7,
                  'rows': [
                      {'col7': '法長ℓ', 'col8': 'ℓ≧5m', 'col9': '法長の－4％', 'col10': None},
                  ]})

    # Case 6: 路体盛土工路床盛土工 条34枝番1
    cases.append({'jou': '３ ４', 'eda': '１', 'koushu': '路体盛土工路床盛土工',
                  'section_hint': '道 路', 'pdf_page': 9,
                  'rows': [
                      {'col7': '法長ℓ', 'col8': 'ℓ≧5m', 'col9': '法長の－4％', 'col10': None},
                  ]})

    # Case 7: 小型標識工 条6
    cases.append({'jou': '６', 'eda': '２', 'koushu': '小型標識工',
                  'section_hint': '共 通 的 工 種', 'pdf_page': 12,
                  'rows': [
                      {'col7': '設置高さＨ', 'col8': None, 'col9': '設計値以上', 'col10': None},
                      {'col7': '根入れ長', 'col8': None, 'col9': '設計値以上', 'col10': None},
                  ]})

    # Case 8: 防止柵工 条7枝番2 → 高さｈ
    cases.append({'jou': '７', 'eda': '２', 'koushu': '防止柵工',
                  'section_hint': '共 通 的 工 種', 'pdf_page': 13,
                  'rows': [
                      {'col7': '高さｈ', 'col8': None, 'col9': '設計値以上', 'col10': None},
                  ]})

    # Case 9: 路側防護柵工（ガードレール）条8枝番1
    cases.append({'jou': '８', 'eda': '１', 'koushu': '路側防護柵工（ガードレール）',
                  'section_hint': '共 通 的 工 種', 'pdf_page': 13,
                  'rows': [
                      {'col7': '基準高▽', 'col8': None, 'col9': '±30', 'col10': None},
                  ]})

    # Case 10: 路側防護柵工（ガードケーブル）条8枝番2
    cases.append({'jou': '８', 'eda': '２', 'koushu': '路側防護柵工（ガードケーブル）',
                  'section_hint': '共 通 的 工 種', 'pdf_page': 14,
                  'rows': [
                      {'col7': 'ケーブル取付高Ｈ', 'col8': None, 'col9': '＋30 －20', 'col10': None},
                      {'col7': '基礎 延長Ｌ', 'col8': None, 'col9': '－100', 'col10': None},
                  ]})

    # Case 11: 根固めブロック工 条17枝番2
    cases.append({'jou': '17', 'eda': '２', 'koushu': '根固めブロック工',
                  'section_hint': '共 通 的 工 種', 'pdf_page': 20,
                  'rows': [
                      {'col7': '層積 厚さｔ', 'col8': None, 'col9': '－20', 'col10': None},
                      {'col7': '層積 幅ｗ1,ｗ2', 'col8': None, 'col9': '－20', 'col10': None},
                      {'col7': '層積 延長Ｌ1,Ｌ2', 'col8': None, 'col9': '－200', 'col10': None},
                      {'col7': '乱積 厚さｔ', 'col8': None, 'col9': '±ｔ／2', 'col10': None},
                  ]})

    # Case 12: 伸縮装置工（鋼製フィンガージョイント）条24枝番2
    cases.append({'jou': '24', 'eda': '２', 'koushu': '伸縮装置工（鋼製フィンガージョイント）',
                  'section_hint': None, 'pdf_page': 22,
                  'rows': [
                      {'col7': '遊間ℓ', 'col8': None, 'col9': '±10', 'col10': None},
                  ]})

    # Case 13: 羽口工（じゃかご）条27枝番1
    cases.append({'jou': '27', 'eda': '１', 'koushu': '羽口工（じゃかご）',
                  'section_hint': None, 'pdf_page': 23,
                  'rows': [
                      {'col7': '延長Ｌ', 'col8': None, 'col9': '－200', 'col10': None},
                  ]})

    # Case 14: コンクリートブロック工 条3枝番1
    cases.append({'jou': '３', 'eda': '１', 'koushu': 'コンクリートブロック工',
                  'section_hint': None, 'pdf_page': 30,
                  'rows': [
                      {'col7': '法長ℓ', 'col8': 'ℓ≧3m', 'col9': '法長の－4％', 'col10': None},
                  ]})

    # Case 15: 緑化ブロック工 条4
    cases.append({'jou': '４', 'eda': '', 'koushu': '緑化ブロック工',
                  'section_hint': None, 'pdf_page': 31,
                  'rows': [
                      {'col7': '法長ℓ', 'col8': 'ℓ≧3m', 'col9': '法長の－4％', 'col10': None},
                  ]})

    # Case 16: 石積（張）工 条5
    cases.append({'jou': '５', 'eda': '', 'koushu': '石積（張）工',
                  'section_hint': None, 'pdf_page': 31,
                  'rows': [
                      {'col7': '法長ℓ', 'col8': 'ℓ≧3m', 'col9': '法長の－4％', 'col10': None},
                  ]})

    # Case 17: 橋面防水工（シート系床版防水層）条6枝番4
    cases.append({'jou': '６', 'eda': '４', 'koushu': '橋面防水工（シート系床版防水層）',
                  'section_hint': None, 'pdf_page': 32,
                  'rows': [
                      {'col7': 'シートの重ね幅', 'col8': None, 'col9': '－20～+50', 'col10': None},
                  ]})

    # Cases 18-28: アスファルト舗装工 series
    cases.append({'jou': '７', 'eda': '１', 'koushu': 'アスファルト舗装工（下層路盤工）',
                  'section_hint': None, 'pdf_page': 33,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－100', 'col10': '―'}]})

    cases.append({'jou': '７', 'eda': '２', 'koushu': 'アスファルト舗装工（下層路盤工）（面管理の場合）',
                  'section_hint': None, 'pdf_page': 34,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '＋40 －15'}]})

    cases.append({'jou': '７', 'eda': '３', 'koushu': 'アスファルト舗装工（上層路盤工）粒度調整路盤工',
                  'section_hint': None, 'pdf_page': 35,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－50', 'col10': '―'}]})

    cases.append({'jou': '７', 'eda': '４', 'koushu': 'アスファルト舗装工（上層路盤工）粒度調整路盤工（面管理の場合）',
                  'section_hint': None, 'pdf_page': 36,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－8'}]})

    cases.append({'jou': '７', 'eda': '５', 'koushu': 'アスファルト舗装工（上層路盤工）セメント（石灰）安定処理工',
                  'section_hint': None, 'pdf_page': 37,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－50', 'col10': '―'}]})

    cases.append({'jou': '７', 'eda': '６', 'koushu': 'アスファルト舗装工（上層路盤工）セメント（石灰）安定処理工（面管理の場合）',
                  'section_hint': None, 'pdf_page': 38,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－8'}]})

    cases.append({'jou': '７', 'eda': '７', 'koushu': 'アスファルト舗装工（加熱アスファルト安定処理工）',
                  'section_hint': None, 'pdf_page': 39,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－50', 'col10': '―'}]})

    cases.append({'jou': '７', 'eda': '８', 'koushu': 'アスファルト舗装工（加熱アスファルト安定処理工）（面管理の場合）',
                  'section_hint': None, 'pdf_page': 40,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－5'}]})

    cases.append({'jou': '７', 'eda': '９', 'koushu': 'アスファルト舗装工（基層工）',
                  'section_hint': None, 'pdf_page': 41,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－25', 'col10': '―'}]})

    cases.append({'jou': '７', 'eda': '10', 'koushu': 'アスファルト舗装工（基層工）（面管理の場合）',
                  'section_hint': None, 'pdf_page': 42,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－3'}]})

    cases.append({'jou': '７', 'eda': '11', 'koushu': 'アスファルト舗装工（表層工）',
                  'section_hint': None, 'pdf_page': 43,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－25', 'col10': '―'}]})

    # Cases 29-38: 半たわみ性舗装工 series
    cases.append({'jou': '８', 'eda': '１', 'koushu': '半たわみ性舗装工（下層路盤工）',
                  'section_hint': None, 'pdf_page': 44,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－100', 'col10': '―'}]})

    cases.append({'jou': '８', 'eda': '３', 'koushu': '半たわみ性舗装工（上層路盤工）粒度調整路盤工',
                  'section_hint': None, 'pdf_page': 45,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－50', 'col10': '―'}]})

    cases.append({'jou': '８', 'eda': '４', 'koushu': '半たわみ性舗装工（上層路盤工）粒度調整路盤工（面管理の場合）',
                  'section_hint': None, 'pdf_page': 46,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－8'}]})

    cases.append({'jou': '８', 'eda': '５', 'koushu': '半たわみ性舗装工（上層路盤工）セメント（石灰）安定処理工',
                  'section_hint': None, 'pdf_page': 47,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－50', 'col10': '―'}]})

    cases.append({'jou': '８', 'eda': '６', 'koushu': '半たわみ性舗装工（上層路盤工）セメント（石灰）安定処理工（面管理の場合）',
                  'section_hint': None, 'pdf_page': 48,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－8'}]})

    cases.append({'jou': '８', 'eda': '７', 'koushu': '半たわみ性舗装工（加熱アスファルト安定処理工）',
                  'section_hint': None, 'pdf_page': 49,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－25', 'col10': '―'}]})

    cases.append({'jou': '８', 'eda': '９', 'koushu': '半たわみ性舗装工（基層工）',
                  'section_hint': None, 'pdf_page': 50,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－25', 'col10': '―'}]})

    cases.append({'jou': '８', 'eda': '10', 'koushu': '半たわみ性舗装工（基層工）（面管理の場合）',
                  'section_hint': None, 'pdf_page': 51,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－3'}]})

    cases.append({'jou': '８', 'eda': '11', 'koushu': '半たわみ性舗装工（表層工）',
                  'section_hint': None, 'pdf_page': 52,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－25', 'col10': '―'}]})

    cases.append({'jou': '８', 'eda': '12', 'koushu': '半たわみ性舗装工（表層工）（面管理の場合）',
                  'section_hint': None, 'pdf_page': 53,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－2'}]})

    # Cases 39-49: 排水性舗装工 series
    cases.append({'jou': '９', 'eda': '１', 'koushu': '排水性舗装工（下層路盤工）',
                  'section_hint': None, 'pdf_page': 54,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－100', 'col10': '―'}]})

    cases.append({'jou': '９', 'eda': '２', 'koushu': '排水性舗装工（下層路盤工）（面管理の場合）',
                  'section_hint': None, 'pdf_page': 55,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '＋40 －15'}]})

    cases.append({'jou': '９', 'eda': '３', 'koushu': '排水性舗装工（上層路盤工）粒度調整路盤工',
                  'section_hint': None, 'pdf_page': 56,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－50', 'col10': '―'}]})

    cases.append({'jou': '９', 'eda': '４', 'koushu': '排水性舗装工（上層路盤工）粒度調整路盤工（面管理の場合）',
                  'section_hint': None, 'pdf_page': 57,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－8'}]})

    cases.append({'jou': '９', 'eda': '５', 'koushu': '排水性舗装工（上層路盤工）セメント（石灰）安定処理工',
                  'section_hint': None, 'pdf_page': 58,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－50', 'col10': '―'}]})

    cases.append({'jou': '９', 'eda': '６', 'koushu': '排水性舗装工（上層路盤工）セメント（石灰）安定処理工（面管理の場合）',
                  'section_hint': None, 'pdf_page': 59,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－8'}]})

    cases.append({'jou': '９', 'eda': '７', 'koushu': '排水性舗装工（加熱アスファルト安定処理工）',
                  'section_hint': None, 'pdf_page': 60,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－25', 'col10': '―'}]})

    cases.append({'jou': '９', 'eda': '８', 'koushu': '排水性舗装工（加熱アスファルト安定処理工）（面管理の場合）',
                  'section_hint': None, 'pdf_page': 61,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－5'}]})

    cases.append({'jou': '９', 'eda': '10', 'koushu': '排水性舗装工（基層工）（面管理の場合）',
                  'section_hint': None, 'pdf_page': 62,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－3'}]})

    cases.append({'jou': '９', 'eda': '11', 'koushu': '排水性舗装工（表層工）',
                  'section_hint': None, 'pdf_page': 63,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－25', 'col10': '―'}]})

    cases.append({'jou': '９', 'eda': '12', 'koushu': '排水性舗装工（表層工）（面管理の場合）',
                  'section_hint': None, 'pdf_page': 64,
                  'rows': [
                      {'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－2'},
                      {'col7': '平坦性', 'col8': '―', 'col9': None, 'col10': '3mプロフィルメーター(σ)2.4㎜以下 直読式(足付き)(σ)1.75㎜以下'},
                  ]})

    # Cases 50-53: 透水性舗装工 series
    cases.append({'jou': '10', 'eda': '１', 'koushu': '透水性舗装工（路盤工）',
                  'section_hint': None, 'pdf_page': 65,
                  'rows': [
                      {'col7': '厚さ', 'col8': 'ｔ＜15㎝', 'col9': '設計値以上', 'col10': '－10'},
                      {'col7': '厚さ', 'col8': 'ｔ≧15cm', 'col9': '設計値の－1割', 'col10': '－15'},
                      {'col7': '延長Ｌ', 'col8': None, 'col9': '－200', 'col10': '―'},
                  ]})

    cases.append({'jou': '10', 'eda': '２', 'koushu': '透水性舗装工（路盤工）（面管理の場合）',
                  'section_hint': None, 'pdf_page': 66,
                  'rows': [
                      {'col7': '基準高▽', 'col8': None, 'col9': None, 'col10': '―'},
                      {'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '＋40 －15'},
                      {'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－8'},
                      {'col7': '幅', 'col8': None, 'col9': '－50', 'col10': '―'},
                  ]})

    cases.append({'jou': '10', 'eda': '３', 'koushu': '透水性舗装工（表層工）',
                  'section_hint': None, 'pdf_page': 67,
                  'rows': [
                      {'col7': '幅', 'col8': None, 'col9': '－50', 'col10': '―'},
                      {'col7': '平坦性', 'col8': '―', 'col9': None, 'col10': '3mプロフィルメーター(σ)2.4㎜以下 直読式(足付き)(σ)1.75㎜以下'},
                  ]})

    cases.append({'jou': '10', 'eda': '４', 'koushu': '透水性舗装工（表層工）（面管理の場合）',
                  'section_hint': None, 'pdf_page': 68,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－8'}]})

    # Cases 54-59: グースアスファルト舗装工 series
    cases.append({'jou': '11', 'eda': '１', 'koushu': 'グースアスファルト舗装工（加熱アスファルト安定処理工）',
                  'section_hint': None, 'pdf_page': 69,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－50', 'col10': '―'}]})

    cases.append({'jou': '11', 'eda': '２', 'koushu': 'グースアスファルト舗装工（加熱アスファルト安定処理工）（面管理の場合）',
                  'section_hint': None, 'pdf_page': 70,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－8'}]})

    cases.append({'jou': '11', 'eda': '３', 'koushu': 'グースアスファルト舗装工（基層工）',
                  'section_hint': None, 'pdf_page': 71,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－25', 'col10': '―'}]})

    cases.append({'jou': '11', 'eda': '４', 'koushu': 'グースアスファルト舗装工（基層工）（面管理の場合）',
                  'section_hint': None, 'pdf_page': 72,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－3'}]})

    cases.append({'jou': '11', 'eda': '５', 'koushu': 'グースアスファルト舗装工（表層工）',
                  'section_hint': None, 'pdf_page': 73,
                  'rows': [{'col7': '平坦性', 'col8': '―', 'col9': None, 'col10': '3mプロフィルメーター(σ)2.4㎜以下 直読式(足付き)(σ)1.75㎜以下'}]})

    cases.append({'jou': '11', 'eda': '６', 'koushu': 'グースアスファルト舗装工（表層工）（面管理の場合）',
                  'section_hint': None, 'pdf_page': 74,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－25', 'col10': '―'}]})

    # Cases 60-78: コンクリート舗装工 series
    cases.append({'jou': '12', 'eda': '１', 'koushu': 'コンクリート舗装工（下層路盤工）',
                  'section_hint': None, 'pdf_page': 75,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－100', 'col10': '―'}]})

    cases.append({'jou': '12', 'eda': '２', 'koushu': 'コンクリート舗装工（下層路盤工）（面管理の場合）',
                  'section_hint': None, 'pdf_page': 76,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '＋40 －15'}]})

    cases.append({'jou': '12', 'eda': '３', 'koushu': 'コンクリート舗装工（粒度調整路盤工）',
                  'section_hint': None, 'pdf_page': 77,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－50', 'col10': '―'}]})

    cases.append({'jou': '12', 'eda': '４', 'koushu': 'コンクリート舗装工（粒度調整路盤工）（面管理の場合）',
                  'section_hint': None, 'pdf_page': 78,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－8'}]})

    cases.append({'jou': '12', 'eda': '５', 'koushu': 'コンクリート舗装工（セメント（石灰・瀝青）安定処理工）',
                  'section_hint': None, 'pdf_page': 79,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－50', 'col10': '―'}]})

    cases.append({'jou': '12', 'eda': '６', 'koushu': 'コンクリート舗装工（セメント（石灰・瀝青）安定処理工）（面管理の場合）',
                  'section_hint': None, 'pdf_page': 80,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－8'}]})

    cases.append({'jou': '12', 'eda': '７', 'koushu': 'コンクリート舗装工（アスファルト中間層）',
                  'section_hint': None, 'pdf_page': 81,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－50', 'col10': '―'}]})

    cases.append({'jou': '12', 'eda': '８', 'koushu': 'コンクリート舗装工（アスファルト中間層）（面管理の場合）',
                  'section_hint': None, 'pdf_page': 82,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－5'}]})

    cases.append({'jou': '12', 'eda': '９', 'koushu': 'コンクリート舗装工（コンクリート舗装版工）',
                  'section_hint': None, 'pdf_page': 83,
                  'rows': [{'col7': '平坦性', 'col8': '―', 'col9': None, 'col10': '3mプロフィルメーター(σ)2.4㎜以下 直読式(足付き)(σ)1.75㎜以下'}]})

    cases.append({'jou': '12', 'eda': '10', 'koushu': 'コンクリート舗装工（コンクリート舗装版工）（面管理の場合）',
                  'section_hint': None, 'pdf_page': 84,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－2'}]})

    cases.append({'jou': '12', 'eda': '11', 'koushu': 'コンクリート舗装工（転圧コンクリート版工）下層路盤工',
                  'section_hint': None, 'pdf_page': 85,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－100', 'col10': '―'}]})

    cases.append({'jou': '12', 'eda': '12', 'koushu': 'コンクリート舗装工（転圧コンクリート版工）下層路盤工（面管理の場合）',
                  'section_hint': None, 'pdf_page': 86,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '＋40 －15'}]})

    cases.append({'jou': '12', 'eda': '14', 'koushu': 'コンクリート舗装工（転圧コンクリート版工）粒度調整路盤工（面管理の場合）',
                  'section_hint': None, 'pdf_page': 87,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－8'}]})

    cases.append({'jou': '12', 'eda': '15', 'koushu': 'コンクリート舗装工（転圧コンクリート版工）セメント（石灰・瀝青）安定処理工',
                  'section_hint': None, 'pdf_page': 88,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－50', 'col10': '―'}]})

    cases.append({'jou': '12', 'eda': '16', 'koushu': 'コンクリート舗装工（転圧コンクリート版工）セメント（石灰・瀝青）安定処理工（面管理の場合）',
                  'section_hint': None, 'pdf_page': 89,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－8'}]})

    cases.append({'jou': '12', 'eda': '17', 'koushu': 'コンクリート舗装工（転圧コンクリート版工）アスファルト中間層',
                  'section_hint': None, 'pdf_page': 90,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－50', 'col10': '―'}]})

    cases.append({'jou': '12', 'eda': '18', 'koushu': 'コンクリート舗装工（転圧コンクリート版工）アスファルト中間層（面管理の場合）',
                  'section_hint': None, 'pdf_page': 91,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－3'}]})

    cases.append({'jou': '12', 'eda': '19', 'koushu': 'コンクリート舗装工（転圧コンクリート版工）',
                  'section_hint': None, 'pdf_page': 92,
                  'rows': [{'col7': '平坦性', 'col8': '―', 'col9': None, 'col10': 'コンクリートの硬化後、3mプロフィルメーターにより機械舗設の場合(σ)2.4㎜以下 人力舗設の場合(σ)3㎜以下'}]})

    cases.append({'jou': '12', 'eda': '20', 'koushu': 'コンクリート舗装工（転圧コンクリート版工）（面管理の場合）',
                  'section_hint': None, 'pdf_page': 93,
                  'rows': [{'col7': '厚さあるいは標高較差', 'col8': None, 'col9': None, 'col10': '－3.5'}]})

    # Cases 79-81: 薄層カラー舗装工 series
    cases.append({'jou': '13', 'eda': '２', 'koushu': '薄層カラー舗装工（上層路盤工）粒度調整路盤工',
                  'section_hint': None, 'pdf_page': 94,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－50', 'col10': '―'}]})

    cases.append({'jou': '13', 'eda': '３', 'koushu': '薄層カラー舗装工（上層路盤工）セメント（石灰）安定処理工',
                  'section_hint': None, 'pdf_page': 95,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－50', 'col10': '―'}]})

    cases.append({'jou': '13', 'eda': '４', 'koushu': '薄層カラー舗装工（加熱アスファルト安定処理工）',
                  'section_hint': None, 'pdf_page': 96,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－25', 'col10': '―'}]})

    # Cases 82-84: ブロック舗装工 series
    cases.append({'jou': '14', 'eda': '１', 'koushu': 'ブロック舗装工（下層路盤工）',
                  'section_hint': None, 'pdf_page': 97,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－100', 'col10': '―'}]})

    cases.append({'jou': '14', 'eda': '３', 'koushu': 'ブロック舗装工（上層路盤工）セメント（石灰）安定処理工',
                  'section_hint': None, 'pdf_page': 98,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－50', 'col10': '―'}]})

    cases.append({'jou': '14', 'eda': '４', 'koushu': 'ブロック舗装工（加熱アスファルト安定処理工）',
                  'section_hint': None, 'pdf_page': 99,
                  'rows': [{'col7': '幅', 'col8': None, 'col9': '－25', 'col10': '―'}]})

    # Case 85: 舗装打換え工 条16
    cases.append({'jou': '16', 'eda': '', 'koushu': '舗装打換え工',
                  'section_hint': None, 'pdf_page': 102,
                  'rows': [
                      {'col7': '基準高▽', 'col8': None, 'col9': '±50', 'col10': None},
                      {'col7': '厚さ', 'col8': 'ｔ＜15㎝', 'col9': '設計値以上', 'col10': None},
                      {'col7': '厚さ', 'col8': 'ｔ≧15cm', 'col9': '設計値の－1割', 'col10': None},
                      {'col7': '延長Ｌ', 'col8': None, 'col9': '－200', 'col10': None},
                  ]})

    # Case 86: 地中連続壁工（柱列式）条10
    cases.append({'jou': '10', 'eda': '', 'koushu': '地中連続壁工（柱列式）',
                  'section_hint': None, 'pdf_page': 111,
                  'rows': [
                      # 等厚式
                      {'col7': '深度Ｌ', 'col8': None, 'col9': '設計値以上', 'col10': None,
                       'col6': '地中連続壁工（等厚式）'},
                      {'col7': '配置誤差', 'col8': None, 'col9': '100', 'col10': None,
                       'col6': '地中連続壁工（等厚式）'},
                      {'col7': '鉛直精度', 'col8': None, 'col9': '1／200以内', 'col10': None,
                       'col6': '地中連続壁工（等厚式）'},
                      {'col7': '壁厚ｔ', 'col8': None, 'col9': '設計値以上', 'col10': None,
                       'col6': '地中連続壁工（等厚式）'},
                      # 壁式（原位置土撹拌式）
                      {'col7': '基準高▽', 'col8': None, 'col9': '±50', 'col10': None,
                       'col6': '地中連続壁工（壁式（原位置土撹拌式））'},
                      {'col7': '位置・間隔ｗ', 'col8': None, 'col9': '100', 'col10': None,
                       'col6': '地中連続壁工（壁式（原位置土撹拌式））'},
                      {'col7': '鉛直精度', 'col8': None, 'col9': '1／200以内', 'col10': None,
                       'col6': '地中連続壁工（壁式（原位置土撹拌式））'},
                      {'col7': '壁厚ｔ', 'col8': None, 'col9': '設計値以上', 'col10': None,
                       'col6': '地中連続壁工（壁式（原位置土撹拌式））'},
                      {'col7': '深度Ｌ', 'col8': None, 'col9': '設計値以上', 'col10': None,
                       'col6': '地中連続壁工（壁式（原位置土撹拌式））'},
                      # 壁式（ソイルセメント柱列壁工法等）
                      {'col7': '基準高▽', 'col8': None, 'col9': '±50', 'col10': None,
                       'col6': '地中連続壁工（壁式（ソイルセメント柱列壁工法、等厚式ソイルセメント地中連続壁工法））'},
                      {'col7': '位置・間隔ｗ', 'col8': None, 'col9': '100', 'col10': None,
                       'col6': '地中連続壁工（壁式（ソイルセメント柱列壁工法、等厚式ソイルセメント地中連続壁工法））'},
                      {'col7': '鉛直精度', 'col8': None, 'col9': '1／200以内', 'col10': None,
                       'col6': '地中連続壁工（壁式（ソイルセメント柱列壁工法、等厚式ソイルセメント地中連続壁工法））'},
                      {'col7': '壁厚ｔ', 'col8': None, 'col9': '設計値以上', 'col10': None,
                       'col6': '地中連続壁工（壁式（ソイルセメント柱列壁工法、等厚式ソイルセメント地中連続壁工法））'},
                      {'col7': '深度Ｌ', 'col8': None, 'col9': '設計値以上', 'col10': None,
                       'col6': '地中連続壁工（壁式（ソイルセメント柱列壁工法、等厚式ソイルセメント地中連続壁工法））'},
                      # 壁式（泥水固化処理壁工法、TRD工法）
                      {'col7': '基準高▽', 'col8': None, 'col9': '－50', 'col10': None,
                       'col6': '地中連続壁工（壁式（泥水固化処理壁工法、TRD工法））'},
                      {'col7': '法長ℓ', 'col8': None, 'col9': '－100', 'col10': None,
                       'col6': '地中連続壁工（壁式（泥水固化処理壁工法、TRD工法））'},
                      {'col7': '連壁の長さℓ', 'col8': None, 'col9': '－50', 'col10': None,
                       'col6': '地中連続壁工（壁式（泥水固化処理壁工法、TRD工法））'},
                      {'col7': '変位', 'col8': None, 'col9': '300', 'col10': None,
                       'col6': '地中連続壁工（壁式（泥水固化処理壁工法、TRD工法））'},
                      {'col7': '壁体長Ｌ', 'col8': None, 'col9': '－200', 'col10': None,
                       'col6': '地中連続壁工（壁式（泥水固化処理壁工法、TRD工法））'},
                      # 壁式（場所打ち鉄筋コンクリート壁工法）
                      {'col7': '基準高▽', 'col8': None, 'col9': '±50', 'col10': None,
                       'col6': '地中連続壁工（壁式（場所打ち鉄筋コンクリート壁工法））'},
                      {'col7': '連壁の長さℓ', 'col8': None, 'col9': '－50', 'col10': None,
                       'col6': '地中連続壁工（壁式（場所打ち鉄筋コンクリート壁工法））'},
                      {'col7': '壁体長Ｌ', 'col8': None, 'col9': '－200', 'col10': None,
                       'col6': '地中連続壁工（壁式（場所打ち鉄筋コンクリート壁工法））'},
                  ]})

    # Case 87: 鋳造費（大型ゴム支承工）条1枝番2
    cases.append({'jou': '１', 'eda': '２', 'koushu': '鋳造費（大型ゴム支承工）',
                  'section_hint': None, 'pdf_page': 113,
                  'rows': [
                      {'col7': '厚さｔ', 'col8': '20＜ｔ≦160', 'col9': '±2.5％', 'col10': None},
                      {'col7': '厚さｔ', 'col8': '160＜ｔ', 'col9': '±4', 'col10': None},
                      {'col7': '相対誤差', 'col8': '1,000mm＜ｗ,Ｌ,Ｄ', 'col9': '（ｗ,Ｌ,Ｄ）／1,000', 'col10': None},
                      {'col7': '部材', 'col8': '部材長ℓ（m）', 'col9': '±3… ℓ≦10 ±4… ℓ＞10', 'col10': None},
                      {'col7': '刃口高さｈ（m）', 'col8': None, 'col9': '±2… ｈ≦0.5 ±3… 0.5＜ｈ≦1.0 ±4… 1.0＜ｈ≦2.0', 'col10': None},
                  ]})

    # Case 88: 桁製作工（仮組立シミュレーション①）条3枝番1 page115
    cases.append({'jou': '３', 'eda': '１', 'koushu': '桁製作工（仮組立',
                  'section_hint': None, 'pdf_page': 115,
                  'rows': [
                      {'col7': '全長Ｌ', 'col8': None, 'col9': '±5…Ｌ≦10 ±10…10＜Ｌ≦20 ±（10＋（Ｌ－20）／10）…20＜Ｌ', 'col10': None},
                      {'col7': '支間長Ｌｎ', 'col8': None, 'col9': '±5…Ｌ≦10 ±10…10＜Ｌ≦20 ±（10＋（Ｌ－20）／10）…20＜Ｌ', 'col10': None},
                      {'col7': '有効幅員ｂ', 'col8': None, 'col9': '±5…ｂ≦5 ±（2.5＋ｂ／2）…ｂ＞5', 'col10': None},
                      # 仮組立精度 items (from page 116)
                      {'col7': '主桁・主構の通りδ', 'col8': None, 'col9': '5＋Ｌ／5…Ｌ≦100 25…Ｌ＞100', 'col10': None},
                      {'col7': '主桁・主構のそりδ', 'col8': None, 'col9': '－5～＋5…Ｌ≦20', 'col10': None},
                      {'col7': '橋端における出入差δ', 'col8': None, 'col9': '±10', 'col10': None},
                      {'col7': '鉛直度δ', 'col8': None, 'col9': '3＋ｈ／1,000', 'col10': None},
                      {'col7': '主桁・主構の中心間距離Ｂ', 'col8': None, 'col9': '±4…Ｂ≦2 ±(3＋Ｂ／2)…Ｂ＞2', 'col10': None},
                      {'col7': '主構の組立高さｈ', 'col8': None, 'col9': '±5…ｈ≦5 ±(2.5＋ｈ／2)…ｈ＞5', 'col10': None},
                      {'col7': '圧縮材の曲がりδ', 'col8': None, 'col9': 'ℓ／1,000', 'col10': None},
                      {'col7': 'フランジの直角度δ', 'col8': None, 'col9': 'ｗ／200', 'col10': None},
                  ]})

    # Case 90: 桁製作工（仮組立検査なし）条3枝番2
    cases.append({'jou': '３', 'eda': '２', 'koushu': '桁製作工（仮組立検査を実施しない場合）',
                  'section_hint': None, 'pdf_page': 117,
                  'rows': [
                      {'col7': '部材精度 板の平面度δ', 'col8': '鋼桁等の部材の腹板', 'col9': 'ｈ／250', 'col10': None},
                      {'col7': '部材精度 板の平面度δ', 'col8': '箱桁等のフランジ・鋼床版デッキプレート', 'col9': 'ｂ／150', 'col10': None},
                      {'col7': '部材精度 フランジの直角度δ', 'col8': None, 'col9': 'ｗ／200', 'col10': None},
                      {'col7': '部材精度 部材長ℓ', 'col8': '鋼桁', 'col9': '±3…ℓ≦10 ±4…ℓ＞10', 'col10': None},
                      {'col7': '部材精度 部材長ℓ', 'col8': 'トラス・アーチなど', 'col9': '±2…ℓ≦10 ±3…ℓ＞10', 'col10': None},
                  ]})

    # Case 91: 鋼製伸縮継手製作工 条5
    cases.append({'jou': '５', 'eda': '', 'koushu': '鋼製伸縮継手製作工',
                  'section_hint': None, 'pdf_page': 120,
                  'rows': [{'col7': '遊間ℓ', 'col8': None, 'col9': '±10', 'col10': None}]})

    # Case 92: アンカーフレーム製作工 条8
    cases.append({'jou': '８', 'eda': '', 'koushu': 'アンカーフレーム製作工',
                  'section_hint': None, 'pdf_page': 121,
                  'rows': [
                      {'col7': 'センターボス ボスの高さ', 'col8': None, 'col9': '＋1－0', 'col10': None},
                      {'col7': 'ボス※5 ボスの高さ', 'col8': None, 'col9': '＋1－1', 'col10': None},
                  ]})

    # Case 93: プレビーム用桁製作工 条9
    cases.append({'jou': '９', 'eda': '', 'koushu': 'プレビーム用桁製作工',
                  'section_hint': None, 'pdf_page': 122,
                  'rows': [
                      {'col7': '全移動量ℓ※4', 'col8': 'ℓ＞300mm', 'col9': '±ℓ／100', 'col10': None},
                      {'col7': '普通寸法 ガス切断寸法', 'col8': None, 'col9': 'JIS B0417-1979 B級', 'col10': None},
                  ]})

    # Case 94: 架設工（鋼橋）
    cases.append({'jou': '', 'eda': '', 'koushu': '架設工（鋼橋）',
                  'section_hint': None, 'pdf_page': 124,
                  'rows': [
                      {'col7': '架設時 キャンバー（そり）', 'col8': None, 'col9': 'δ＝δ0＋Ｔｍ×α×Ｌ×10⁶', 'col10': None},
                  ]})

    # Case 95: 植生工（種子散布工等）条2枝番1
    cases.append({'jou': '２', 'eda': '１', 'koushu': '植生工（種子散布工）',
                  'section_hint': None, 'pdf_page': 125,
                  'rows': [
                      {'col7': '厚さｔ', 'col8': None, 'col9': '設計値以上', 'col10': None},
                      {'col7': '法長ℓ', 'col8': None, 'col9': '設計値以上', 'col10': None},
                  ]})

    # Case 96: 植生工（植生基材吹付工）条2枝番2
    cases.append({'jou': '２', 'eda': '２', 'koushu': '植生工（植生基材吹付工）',
                  'section_hint': None, 'pdf_page': 126,
                  'rows': [
                      {'col7': '厚さｔ', 'col8': None, 'col9': '設計値以上', 'col10': None},
                      {'col7': '法長ℓ', 'col8': None, 'col9': '設計値以上', 'col10': None},
                      {'col7': '幅ｗ', 'col8': None, 'col9': '－200', 'col10': None},
                  ]})

    # Case 97: 吹付工（コンクリート）（モルタル）条3
    cases.append({'jou': '３', 'eda': '', 'koushu': '吹付工（コンクリート）（モルタル）',
                  'section_hint': None, 'pdf_page': 127,
                  'rows': [
                      {'col7': '幅ｗ', 'col8': None, 'col9': '－200', 'col10': None},
                      {'col7': '法長ℓ', 'col8': 'ℓ≧5m', 'col9': '法長の－4％', 'col10': None},
                      {'col7': '延長Ｌ', 'col8': None, 'col9': '－200', 'col10': None},
                  ]})

    # Case 98: 法枠工（現場打法枠工）条4枝番1
    cases.append({'jou': '４', 'eda': '１', 'koushu': '法枠工（現場打法枠工）',
                  'section_hint': None, 'pdf_page': 128,
                  'rows': [{'col7': '法長ℓ', 'col8': 'ℓ≧5m', 'col9': '法長の－4％', 'col10': None}]})

    # Case 99: 法枠工（プレキャスト法枠工）条4枝番2
    cases.append({'jou': '４', 'eda': '２', 'koushu': '法枠工（プレキャスト法枠工）',
                  'section_hint': None, 'pdf_page': 128,
                  'rows': [{'col7': '法長ℓ', 'col8': 'ℓ≧5m', 'col9': '法長の－4％', 'col10': None}]})

    # Case 100: 一般事項（場所打擁壁工）条1
    cases.append({'jou': '１', 'eda': '', 'koushu': '一般事項（場所打擁壁工）',
                  'section_hint': None, 'pdf_page': 129,
                  'rows': [{'col7': '法長ℓ', 'col8': 'ℓ≧5m', 'col9': '法長の－4％', 'col10': None}]})

    # Case 101: 補強土壁工 条3
    cases.append({'jou': '３', 'eda': '', 'koushu': '補強土壁工',
                  'section_hint': None, 'pdf_page': 130,
                  'rows': [{'col7': '法長ℓ', 'col8': 'ℓ≧5m', 'col9': '法長の－4％', 'col10': None}]})

    # Case 102: 井桁ブロック工 条4
    cases.append({'jou': '４', 'eda': '', 'koushu': '井桁ブロック工',
                  'section_hint': None, 'pdf_page': 131,
                  'rows': [{'col7': '法長ℓ', 'col8': 'ℓ≧5m', 'col9': '法長の－4％', 'col10': None}]})

    # Case 103: 浚渫船運転工（ポンプ浚渫船）条3枝番1
    cases.append({'jou': '３', 'eda': '１', 'koushu': '浚渫船運転工（ポンプ浚渫船）',
                  'section_hint': None, 'pdf_page': 132,
                  'rows': [
                      {'col7': '基準高▽', 'col8': None, 'col9': '±50', 'col10': None},
                      {'col7': '厚さｔ', 'col8': None, 'col9': '－50', 'col10': None},
                      {'col7': '幅ｗ', 'col8': None, 'col9': '－100', 'col10': None},
                      {'col7': '法長ℓ', 'col8': 'ℓ＜5m', 'col9': '－100', 'col10': None},
                      {'col7': '法長ℓ', 'col8': 'ℓ≧5m', 'col9': '法長の－4％', 'col10': None},
                  ]})

    # Case 104: 床版工 条2
    cases.append({'jou': '２', 'eda': '', 'koushu': '床版工',
                  'section_hint': None, 'pdf_page': 134,
                  'rows': [{'col7': '高さＨ', 'col8': None, 'col9': '±10', 'col10': None}]})

    # Case 105: 堰本体工水叩工土砂吐工 条8910
    cases.append({'jou': '８ ９ 10', 'eda': '', 'koushu': '堰本体工水叩工土砂吐工',
                  'section_hint': None, 'pdf_page': 138,
                  'rows': [{'col7': '延長Ｌ', 'col8': 'Ｌ≧20m', 'col9': '－400', 'col10': None}]})

    # Case 106: 海岸コンクリートブロック工 (港湾) 条4
    cases.append({'jou': '４', 'eda': '', 'koushu': '海岸コンクリートブロック工',
                  'section_hint': '港 湾', 'pdf_page': 143,
                  'rows': [{'col7': '延長Ｌ', 'col8': None, 'col9': '－200', 'col10': None}]})

    # Case 107: コンクリート被覆工 条5
    cases.append({'jou': '５', 'eda': '', 'koushu': 'コンクリート被覆工',
                  'section_hint': None, 'pdf_page': 144,
                  'rows': [
                      {'col7': '法長ℓ', 'col8': 'ℓ＜3m', 'col9': '－50', 'col10': None},
                      {'col7': '法長ℓ', 'col8': 'ℓ≧3m', 'col9': '法長の－4％', 'col10': None},
                  ]})

    # Case 108: 捨石工 (港湾) 条4
    cases.append({'jou': '４', 'eda': '', 'koushu': '捨石工',
                  'section_hint': '港 湾', 'pdf_page': 145,
                  'rows': [
                      {'col7': '天端幅ｗ', 'col8': None, 'col9': '－100', 'col10': None},
                      {'col7': '法長ℓ', 'col8': 'ℓ＜5m', 'col9': '－100', 'col10': None},
                      {'col7': '法長ℓ', 'col8': 'ℓ≧5m', 'col9': '法長の－4％', 'col10': None},
                      {'col7': '延長Ｌ', 'col8': 'Ｌ＜20m', 'col9': '－50', 'col10': None},
                      {'col7': '延長Ｌ', 'col8': 'Ｌ≧20m', 'col9': '－100', 'col10': None},
                  ]})

    # Case 109: 捨石工 条2 (海岸)
    cases.append({'jou': '２', 'eda': '', 'koushu': '捨石工',
                  'section_hint': '海 岸', 'pdf_page': 146,
                  'rows': [{'col7': '延長Ｌ', 'col8': None, 'col9': '－200', 'col10': None}]})

    # Case 110: 海岸コンクリートブロック工 条5 (海岸)
    cases.append({'jou': '５', 'eda': '', 'koushu': '海岸コンクリートブロック工',
                  'section_hint': '海 岸', 'pdf_page': 146,
                  'rows': [
                      {'col7': '厚さｔ，ｔ₁，ｔ₂', 'col8': None, 'col9': '－20', 'col10': None},
                      {'col7': '延長Ｌ', 'col8': None, 'col9': '－200', 'col10': None},
                  ]})

    # Case 111: 石枠工 条9
    cases.append({'jou': '９', 'eda': '', 'koushu': '石枠工',
                  'section_hint': None, 'pdf_page': 147,
                  'rows': [{'col7': '高さｈ，ｈ₁，ｈ₂', 'col8': 'ｈ≧3m', 'col9': '－50', 'col10': None}]})

    # Case 112: ケーソン工（ケーソン工製作）条11枝番1
    cases.append({'jou': '11', 'eda': '１', 'koushu': 'ケーソン工（ケーソン工製作）',
                  'section_hint': None, 'pdf_page': 148,
                  'rows': [{'col7': '壁厚ｔ', 'col8': None, 'col9': '－20', 'col10': None}]})

    # Case 113: ケーソン工（ケーソン工据付）条11枝番2
    cases.append({'jou': '11', 'eda': '２', 'koushu': 'ケーソン工（ケーソン工据付）',
                  'section_hint': None, 'pdf_page': 148,
                  'rows': [
                      {'col7': '偏心量ｄ', 'col8': None, 'col9': '300以内', 'col10': None},
                      {'col7': '基準高▽', 'col8': None, 'col9': '±100', 'col10': None},
                  ]})

    # Case 114: ケーソン工（突堤上部工）条11枝番3
    cases.append({'jou': '11', 'eda': '３', 'koushu': 'ケーソン工（突堤上部工）',
                  'section_hint': None, 'pdf_page': 149,
                  'rows': [{'col7': '延長Ｌ', 'col8': None, 'col9': '－200', 'col10': None}]})

    # Case 115: セルラー工（突堤上部工）条12枝番3
    cases.append({'jou': '12', 'eda': '３', 'koushu': 'セルラー工（突堤上部工）',
                  'section_hint': None, 'pdf_page': 149,
                  'rows': [{'col7': '延長Ｌ', 'col8': None, 'col9': '－200', 'col10': None}]})

    # Case 116: 捨石工 条2 (漁港)
    cases.append({'jou': '２', 'eda': '', 'koushu': '捨石工',
                  'section_hint': '漁 港', 'pdf_page': 150,
                  'rows': [{'col7': '基準高▽', 'col8': None, 'col9': '±50', 'col10': None}]})

    # Case 117: 根固めブロック工 条3 (漁港)
    cases.append({'jou': '３', 'eda': '', 'koushu': '根固めブロック工',
                  'section_hint': '漁 港', 'pdf_page': 150,
                  'rows': [
                      {'col7': '基準高▽', 'col8': '乱積', 'col9': '±ｔ／2', 'col10': None},
                      {'col7': '厚さｔ', 'col8': '乱積', 'col9': '－ｔ／2', 'col10': None},
                      {'col7': '幅ｗ', 'col8': '乱積', 'col9': '－ｔ／2', 'col10': None},
                  ]})

    # Case 118: 消波ブロック工 条3 (漁港)
    cases.append({'jou': '３', 'eda': '', 'koushu': '消波ブロック工',
                  'section_hint': '漁 港', 'pdf_page': 150,
                  'rows': [{'col7': '基準高▽', 'col8': '乱積', 'col9': '±ブロックの高さの1／2', 'col10': None}]})

    # Case 119: 捨石工 条3 (漁港・別ブロック)
    cases.append({'jou': '３', 'eda': '', 'koushu': '捨石工',
                  'section_hint': '漁 港', 'pdf_page': 151,
                  'rows': [
                      {'col7': '天端幅ｗ', 'col8': None, 'col9': '－100', 'col10': None},
                      {'col7': '法長ℓ', 'col8': 'ℓ＜5m', 'col9': '－100', 'col10': None},
                      {'col7': '法長ℓ', 'col8': 'ℓ≧5m', 'col9': '法長の－4％', 'col10': None},
                      {'col7': '延長Ｌ', 'col8': None, 'col9': '－200', 'col10': None},
                  ]})

    # Case 120: 鋼製堰堤本体工（不透過型）条5枝番1
    cases.append({'jou': '５', 'eda': '１', 'koushu': '鋼製堰堤本体工（不透過型）',
                  'section_hint': None, 'pdf_page': 153,
                  'rows': [
                      {'col7': '基準高▽', 'col8': None, 'col9': '±30', 'col10': None,
                       'col6': 'コンクリート堰堤工'},
                      {'col7': '幅ｗ', 'col8': None, 'col9': '－100', 'col10': None,
                       'col6': 'コンクリート堰堤工'},
                      {'col7': '延長Ｌ', 'col8': None, 'col9': '－100', 'col10': None,
                       'col6': 'コンクリート堰堤工'},
                      {'col7': '堤高▽', 'col8': None, 'col9': '±50', 'col10': None},
                      {'col7': '長さθ₁，θ₂', 'col8': '水通し部', 'col9': '±100', 'col10': None},
                      {'col7': '幅ｗ₁，ｗ₃', 'col8': '水通し部', 'col9': '±50', 'col10': None},
                      {'col7': '下流側倒れ△', 'col8': '水通し部', 'col9': '±0.02Ｈ₁', 'col10': None},
                      {'col7': '袖高▽', 'col8': '袖部', 'col9': '±50', 'col10': None},
                      {'col7': '幅ｗ₂', 'col8': '袖部', 'col9': '±50', 'col10': None},
                      {'col7': '下流側倒れ△', 'col8': '袖部', 'col9': '±0.02Ｈ₂', 'col10': None},
                      {'col7': '厚さｔ', 'col8': 'ｔ≧100', 'col9': '設計値の－2割', 'col10': None},
                  ]})

    # Case 121: 場所打函渠工 条6
    cases.append({'jou': '６', 'eda': '', 'koushu': '場所打函渠工',
                  'section_hint': None, 'pdf_page': 163,
                  'rows': [{'col7': '延長Ｌ', 'col8': None, 'col9': '－200', 'col10': None}]})

    # Case 122: 防雪柵工 条6
    cases.append({'jou': '６', 'eda': '', 'koushu': '防雪柵工',
                  'section_hint': None, 'pdf_page': 164,
                  'rows': [{'col7': '延長Ｌ', 'col8': None, 'col9': '－200', 'col10': None}]})

    # Case 123: 雪崩予防柵工 条7
    cases.append({'jou': '７', 'eda': '', 'koushu': '雪崩予防柵工',
                  'section_hint': None, 'pdf_page': 165,
                  'rows': [
                      {'col7': '基礎 幅ｗ₁，ｗ₂', 'col8': None, 'col9': '－30', 'col10': None},
                      {'col7': '基礎 高さｈ', 'col8': None, 'col9': '－30', 'col10': None},
                  ]})

    # Case 124: 遮音壁本体工 条5
    cases.append({'jou': '５', 'eda': '', 'koushu': '遮音壁本体工',
                  'section_hint': None, 'pdf_page': 165,
                  'rows': [
                      {'col7': '間隔ｗ₁，ｗ₂', 'col8': None, 'col9': '±15', 'col10': None},
                      {'col7': 'ずれａ', 'col8': '支柱', 'col9': '10', 'col10': None},
                      {'col7': 'ねじれｂ－ｃ', 'col8': '支柱', 'col9': '5', 'col10': None},
                      {'col7': '倒れｄ', 'col8': '支柱', 'col9': 'ｈ×0.5%', 'col10': None},
                  ]})

    # Case 125: 歩道路盤工
    cases.append({'jou': '', 'eda': '', 'koushu': '歩道路盤工',
                  'section_hint': None, 'pdf_page': 166,
                  'rows': [
                      {'col7': '厚さ', 'col8': 'ｔ＜15㎝', 'col9': '－30', 'col10': '－10'},
                      {'col7': '厚さ', 'col8': 'ｔ≧15cm', 'col9': '－45', 'col10': '－15'},
                  ]})

    # Case 126: 踏掛版工（コンクリート工）条4
    cases.append({'jou': '４', 'eda': '', 'koushu': '踏掛版工（コンクリート工）',
                  'section_hint': None, 'pdf_page': 168,
                  'rows': [
                      {'col7': '基準高', 'col8': None, 'col9': '±20', 'col10': None},
                      {'col7': '各部の厚さ', 'col8': None, 'col9': '±20', 'col10': None},
                      {'col7': '各部の長さ', 'col8': None, 'col9': '±30', 'col10': None},
                      {'col7': '各部の長さ', 'col8': 'ラバーシュー', 'col9': '±20', 'col10': None,
                       'col6': '踏掛版工（ラバーシュー）'},
                  ]})

    # Case 127: 鋼製橋脚製作工 条3
    cases.append({'jou': '３', 'eda': '', 'koushu': '鋼製橋脚製作工',
                  'section_hint': None, 'pdf_page': 171,
                  'rows': [
                      {'col7': '脚柱とベースプレートの鉛直度δ', 'col8': None, 'col9': 'ｗ／500', 'col10': None},
                      {'col7': 'ベースプレート 孔の位置', 'col8': '部材', 'col9': '±2', 'col10': None},
                      {'col7': 'ベースプレート 孔の径ｄ', 'col8': '部材', 'col9': '0～5', 'col10': None},
                      {'col7': '柱の中心間隔、対角長Ｌ（m）', 'col8': None, 'col9': '±5…Ｌ≦10m ±10…10＜Ｌ≦20m ±(10＋(Ｌ－20)／10)…20m＜Ｌ', 'col10': None},
                      {'col7': 'はりのキャンバー及び柱の曲がりδ', 'col8': '仮組立時', 'col9': 'Ｌ／1,000', 'col10': None},
                      {'col7': '柱の鉛直度δ', 'col8': None, 'col9': '10…Ｈ≦10 Ｈ…Ｈ＞10', 'col10': None},
                      {'col7': '部材精度 板の平面度δ', 'col8': '鋼桁及びトラス等の部材の腹板', 'col9': 'ｈ／250', 'col10': None},
                      {'col7': '部材精度 板の平面度δ', 'col8': '箱桁及びトラス等のフランジ・鋼床版デッキプレート', 'col9': 'ｂ／150', 'col10': None},
                      {'col7': '部材精度 フランジの直角度δ', 'col8': None, 'col9': 'ｗ／200', 'col10': None},
                      {'col7': '部材精度 圧縮材の曲がりδ', 'col8': None, 'col9': 'ℓ／1,000', 'col10': None},
                  ]})

    # Case 128: 支承工（鋼製支承）条10枝番1
    cases.append({'jou': '10', 'eda': '１', 'koushu': '支承工（鋼製支承）',
                  'section_hint': None, 'pdf_page': 180,
                  'rows': [
                      {'col7': '据付け高さ', 'col8': '注1）', 'col9': '±5', 'col10': None},
                      {'col7': '可動支承の移動可能量', 'col8': '注2）', 'col9': '設計移動量以上', 'col10': None},
                      {'col7': '支承中心間隔（橋軸直角方向）', 'col8': 'コンクリート橋', 'col9': '±5', 'col10': None},
                      {'col7': '支承中心間隔（橋軸直角方向）', 'col8': '鋼橋', 'col9': '±(4＋0.5×(Ｂ－2))', 'col10': None},
                      {'col7': '水平度 橋軸方向', 'col8': None, 'col9': '1／100', 'col10': None},
                      {'col7': '水平度 橋軸直角方向', 'col8': None, 'col9': '1／100', 'col10': None},
                      {'col7': '可動支承の橋軸方向のずれ 同一支承線上の相対誤差', 'col8': None, 'col9': '5', 'col10': None},
                  ]})

    # Case 129: 支承工（ゴム支承）条10枝番2
    cases.append({'jou': '10', 'eda': '２', 'koushu': '支承工（ゴム支承）',
                  'section_hint': None, 'pdf_page': 180,
                  'rows': [
                      {'col7': '据付け高さ', 'col8': '注1）', 'col9': '±5', 'col10': None},
                      {'col7': '可動支承の移動可能量', 'col8': '注2）', 'col9': '設計移動量以上', 'col10': None},
                  ]})

    # Case 130: 坑門本体工 条4
    cases.append({'jou': '４', 'eda': '', 'koushu': '坑門本体工',
                  'section_hint': None, 'pdf_page': 185,
                  'rows': [{'col7': '高さｈ', 'col8': 'ｈ≧3m', 'col9': '－100', 'col10': None}]})

    # Case 131: 路上再生工 条7
    cases.append({'jou': '７', 'eda': '', 'koushu': '路上再生工',
                  'section_hint': None, 'pdf_page': 191,
                  'rows': [
                      {'col7': '幅ｗ', 'col8': '路盤工', 'col9': '－50', 'col10': None},
                      {'col7': '延長Ｌ', 'col8': None, 'col9': '－100', 'col10': None},
                  ]})

    return cases


def normalize(s):
    """比較用にテキストを正規化する"""
    if s is None:
        return ''
    return str(s).strip().replace('\u3000', ' ')


def find_block(ws, case, max_row):
    """ケースに該当するDB内ブロック（行範囲）を検索する。
    Returns: (start_row, end_row) or None
    """
    koushu_query = case['koushu']
    section_hint = case.get('section_hint')

    candidates = []
    for r in range(2, max_row + 1):
        koushu_val = normalize(ws.cell(r, 6).value)
        if not koushu_val:
            continue
        if koushu_query in koushu_val or koushu_val in koushu_query:
            # section_hintで絞り込み
            if section_hint:
                setsu_val = normalize(ws.cell(r, 3).value)
                if section_hint not in setsu_val:
                    continue
            candidates.append(r)

    if not candidates:
        return None

    # 連続ブロックをグループ化
    blocks = []
    current_block = [candidates[0]]
    for i in range(1, len(candidates)):
        if candidates[i] - candidates[i - 1] <= 5:  # 5行以内なら同一ブロック
            current_block.append(candidates[i])
        else:
            blocks.append(current_block)
            current_block = [candidates[i]]
    blocks.append(current_block)

    # 複数ブロックがある場合、条・枝番で絞り込み
    if len(blocks) > 1 and case.get('jou'):
        filtered = []
        for block in blocks:
            jou_val = normalize(ws.cell(block[0], 4).value)
            if case['jou'] in jou_val or jou_val in case['jou']:
                filtered.append(block)
        if filtered:
            blocks = filtered

    if not blocks:
        return None

    # 最初のマッチブロックを使用
    block = blocks[0]
    start = block[0]
    end = block[-1]

    # ブロック末尾を拡張（同じ条・枝番の行が続く場合）
    while end + 1 <= max_row:
        next_koushu = normalize(ws.cell(end + 1, 6).value)
        if next_koushu and (koushu_query in next_koushu or next_koushu in koushu_query):
            end += 1
        else:
            break

    return (start, end)


def row_exists(ws, start, end, new_row):
    """ブロック内に同じ測定項目+規格値_条件+規格値の行が既に存在するか"""
    new_item = normalize(new_row.get('col7', ''))
    new_cond = normalize(new_row.get('col8', ''))
    new_kika = normalize(new_row.get('col9', ''))
    new_kika_indiv = normalize(new_row.get('col10', ''))

    # 工種が異なる新行の場合（col6指定あり）、その工種が既にあるかもチェック
    new_koushu = new_row.get('col6')

    for r in range(start, end + 1):
        item = normalize(ws.cell(r, 7).value)
        cond = normalize(ws.cell(r, 8).value)
        kika = normalize(ws.cell(r, 9).value)
        kika_indiv = normalize(ws.cell(r, 10).value)
        koushu = normalize(ws.cell(r, 6).value)

        # 新しい工種名が指定されている場合、工種名もチェック
        if new_koushu:
            if normalize(new_koushu) not in koushu and koushu not in normalize(new_koushu):
                continue

        # 測定項目の一致チェック（部分一致）
        if new_item and item:
            if new_item not in item and item not in new_item:
                continue
        elif new_item != item:
            continue

        # 規格値_条件のチェック（空同士もOK）
        if new_cond and cond:
            if new_cond not in cond and cond not in new_cond:
                continue
        elif new_cond != cond:
            # 片方だけ空の場合はスキップしない（条件なし同士はOK）
            if new_cond or cond:
                continue

        # 規格値もチェック
        if new_kika and kika:
            if new_kika in kika or kika in new_kika:
                return True
        elif not new_kika and not kika:
            # 規格値_個々もチェック
            if new_kika_indiv and kika_indiv:
                if new_kika_indiv in kika_indiv or kika_indiv in new_kika_indiv:
                    return True
            return True
        elif new_kika == kika:
            return True

    return False


def insert_row(ws, insert_at, ref_row, row_data):
    """指定位置に新行を挿入する"""
    ws.insert_rows(insert_at)

    # 参照行からコピーするカラム: 編,章,節,条,枝番,工種,測定基準,測定箇所,摘要
    copy_cols = [1, 2, 3, 4, 5, 6, 11, 12, 13]
    for col in copy_cols:
        src_cell = ws.cell(ref_row if ref_row < insert_at else ref_row + 1, col)
        dst_cell = ws.cell(insert_at, col)
        dst_cell.value = src_cell.value
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.border = copy(src_cell.border)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format

    # 明示指定カラムを上書き
    col_map = {'col6': 6, 'col7': 7, 'col8': 8, 'col9': 9, 'col10': 10}
    for key, col_idx in col_map.items():
        if key in row_data:
            ws.cell(insert_at, col_idx).value = row_data[key]

    # データカラムのスタイルもコピー
    ref_actual = ref_row if ref_row < insert_at else ref_row + 1
    for col in [7, 8, 9, 10]:
        src_cell = ws.cell(ref_actual, col)
        dst_cell = ws.cell(insert_at, col)
        if src_cell.has_style:
            dst_cell.font = copy(src_cell.font)
            dst_cell.fill = copy(src_cell.fill)
            dst_cell.border = copy(src_cell.border)
            dst_cell.alignment = copy(src_cell.alignment)
            dst_cell.number_format = src_cell.number_format


def apply_header_style(ws):
    """ヘッダー行に青背景・白太字スタイルを適用"""
    header_font = Font(name='ＭＳ Ｐゴシック', size=10, bold=True, color='FFFFFFFF')
    header_fill = PatternFill(patternType='solid', fgColor='FF1F4E79')
    for col in range(1, 14):
        cell = ws.cell(1, col)
        cell.font = header_font
        cell.fill = header_fill


def update_version(wb):
    """バージョン情報シートの作成日時を更新"""
    vs = wb[VERSION_SHEET]
    vs.cell(2, 4).value = datetime.now().strftime('%Y-%m-%d %H:%M:%S')


def main():
    print("出来形管理DB修正スクリプト（動的マッチング版）")
    print("=" * 60)

    wb = openpyxl.load_workbook(DB_PATH)
    ws = wb[SHEET_NAME]
    original_rows = ws.max_row - 1
    print(f"現在の行数: {original_rows}")

    cases = define_cases()
    print(f"チェック対象: {len(cases)} ケース")

    total_inserted = 0
    not_found = []
    already_exists = []

    # 各ケースを逆順に処理（後ろから挿入してインデックスずれ防止）
    for ci in range(len(cases) - 1, -1, -1):
        case = cases[ci]
        case_num = ci + 1
        max_row = ws.max_row

        block = find_block(ws, case, max_row)
        if not block:
            not_found.append(f"Case {case_num}: {case['koushu']}")
            continue

        start, end = block

        # ブロック内で欠落行を特定
        rows_to_insert = []
        for row_data in case['rows']:
            if not row_exists(ws, start, end, row_data):
                rows_to_insert.append(row_data)
            else:
                already_exists.append(f"Case {case_num}: {row_data.get('col7', '')} ({case['koushu']})")

        if not rows_to_insert:
            continue

        # ブロック末尾に挿入（逆順で挿入して順序を保持）
        for row_data in reversed(rows_to_insert):
            insert_at = end + 1
            insert_row(ws, insert_at, end, row_data)
            total_inserted += 1
            print(f"  Case {case_num}: 挿入 [{row_data.get('col7', '')}] → 行{insert_at} ({case['koushu']})")

    print(f"\n{'=' * 60}")
    print(f"挿入完了: {total_inserted} 行追加")

    if not_found:
        print(f"\n⚠ ブロック未発見 ({len(not_found)}件):")
        for nf in not_found:
            print(f"  {nf}")

    if already_exists:
        print(f"\n✓ 既に存在 ({len(already_exists)}件): スキップ済み")

    # ヘッダースタイル
    apply_header_style(ws)

    # バージョン情報更新
    update_version(wb)

    # 保存
    new_rows = ws.max_row - 1
    print(f"\n保存中: {DB_PATH}")
    wb.save(DB_PATH)
    print(f"完了! {original_rows} → {new_rows} 行 （+{new_rows - original_rows}行）")


if __name__ == '__main__':
    main()
