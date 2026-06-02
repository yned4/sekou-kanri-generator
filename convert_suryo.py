"""
convert_suryo.py
数量総括表PDFをExcelに変換するスクリプト。

使い方:
    python convert_suryo.py <数量総括表.pdf> [出力先.xlsx]

出力先省略時は入力ファイルと同じディレクトリに保存する。
"""

import sys
import re
import io
from pathlib import Path
from typing import Optional

import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ===== 定数 =====
VALID_COL_COUNTS = {22, 23}

# コスト集計項目（工種ではない）
COST_WORDS = {
    "直接工事費", "共通仮設費", "共通仮設費（率計上）", "純工事費",
    "現場管理費", "工事原価", "一般管理費等", "工事価格",
    "消費税相当額", "工事費計", "運搬費", "重建設機械分解組立輸送費",
    "技術管理費", "現場環境改善費（率計上）",
    "ｼｽﾃﾑ初期費(ICT)", "システム初期費(ICT)",
}

# 列インデックス（列数によって異なる）
COL_KOJYO = 1   # 工種/種別/細別/名称（テーブル抽出用）
COL_CONFIG = {
    23: {"規格": 5, "単位": 9,  "数量": 14},
    22: {"規格": 5, "単位": 8,  "数量": 13},
}

# 工種エリアのx座標範囲（ページ左端から工種テキストが入る範囲）
KOJYO_X_MIN = 35
KOJYO_X_MAX = 230

# 階層列名（最大4レベル）
LEVEL_COLS = ["工種", "種別", "細別", "名称"]

# 出力列
OUT_COLS = ["工事名", "工種", "種別", "細別", "名称", "規格", "単位", "数量（今回）"]

# スタイル定数
HEADER_BG  = "1F4E79"
HEADER_FG  = "FFFFFF"
SECTION_BG = "D6E4F0"
MIN_WIDTH  = 8
MAX_WIDTH  = 60


def _clean(val) -> str:
    """None/改行を正規化する。"""
    if val is None:
        return ""
    s = str(val).replace("\n", " ").strip()
    return re.sub(r"\s+", " ", s)


def _is_skip_row(val: str) -> bool:
    """スキップすべき行か判定する。"""
    if not val:
        return True
    if val.startswith("(") or val.startswith("（"):
        return True
    if val in COST_WORDS:
        return True
    if "工事区分" in val or "工事名" in val:
        return True
    return False


# ---------------------------------------------------------------------------
# x座標ベースの階層判定
# ---------------------------------------------------------------------------

def _build_text_x0_map(page) -> dict[str, float]:
    """
    工種エリア（KOJYO_X_MIN〜KOJYO_X_MAX）内のテキストを
    行ごとにまとめ、{テキスト: x0} のマッピングを返す。

    同一テキストが複数回出現する場合は最初の出現のx0を使用する。
    （同じ品目は常に同じ階層レベルに配置されるため）
    """
    words = page.extract_words(x_tolerance=3, y_tolerance=3)
    area_words = [w for w in words
                  if KOJYO_X_MIN <= w["x0"] < KOJYO_X_MAX and w["text"].strip()]

    # y座標でグループ化（5pt以内を同一行とみなす）
    rows: dict[float, list] = {}
    for w in area_words:
        placed = False
        for ky in rows:
            if abs(ky - w["top"]) < 5:
                rows[ky].append(w)
                placed = True
                break
        if not placed:
            rows[w["top"]] = [w]

    result: dict[str, float] = {}
    for ws in rows.values():
        ws_sorted = sorted(ws, key=lambda w: w["x0"])
        text = _clean(" ".join(w["text"] for w in ws_sorted))
        x0 = ws_sorted[0]["x0"]
        if text and text not in result:
            result[text] = x0

    return result


def _x0_to_level(x0_vals: list[float]) -> dict[float, int]:
    """
    x0座標の一覧からレベルマッピングを構築する。
    x0を5pt単位に丸めて重複を除去し、昇順にレベル番号を割り当てる。
    """
    rounded = sorted(set(round(x / 5) * 5 for x in x0_vals))
    return {r: i for i, r in enumerate(rounded)}


# ---------------------------------------------------------------------------
# ページ単位の抽出
# ---------------------------------------------------------------------------

def _extract_page(page) -> list[dict]:
    """1ページ分を処理してレコードリストを返す。"""
    tables = page.extract_tables()
    valid = [t for t in tables if t and len(t[0]) in VALID_COL_COUNTS]
    if not valid:
        return []

    table = valid[0]
    ncol = len(table[0])
    cfg  = COL_CONFIG[ncol]

    # 工事名
    kojyo_mei = ""
    for cell in table[0]:
        v = _clean(cell)
        if v and "令和" in v:
            kojyo_mei = v
            break

    # テキスト → x0 のマッピング
    text_to_x0 = _build_text_x0_map(page)

    records = []
    for row in table[2:]:
        val = _clean(row[COL_KOJYO])
        if _is_skip_row(val):
            continue

        kiku  = _clean(row[cfg["規格"]]) if cfg["規格"] < len(row) else ""
        tani  = _clean(row[cfg["単位"]]) if cfg["単位"] < len(row) else ""
        suryo = _clean(row[cfg["数量"]]) if cfg["数量"] < len(row) else ""

        records.append({
            "工事名":      kojyo_mei,
            "_val":        val,
            "_x0":         text_to_x0.get(val),   # None の場合はレベル不明
            "規格":        kiku,
            "単位":        tani,
            "数量（今回）": suryo,
        })

    return records


# ---------------------------------------------------------------------------
# 階層列への変換
# ---------------------------------------------------------------------------

def _build_hierarchy(records: list[dict]) -> pd.DataFrame:
    """
    各レコードの _x0 からレベルを判定して LEVEL_COLS に展開する。

    x0 が取れなかった行（_x0=None）は前のレコードと同じレベルを使う。
    """
    x0_vals = [r["_x0"] for r in records if r["_x0"] is not None]
    if x0_vals:
        level_map = _x0_to_level(x0_vals)
        base_rounded = min(level_map)
    else:
        level_map = {}
        base_rounded = 0

    n_levels = len(LEVEL_COLS)
    current = [""] * n_levels
    prev_level = 0

    rows = []
    for r in records:
        x0 = r["_x0"]
        if x0 is not None:
            rounded = round(x0 / 5) * 5
            level = level_map.get(rounded, 0)
        else:
            level = prev_level  # x0不明は前行と同レベルとみなす

        level = max(0, min(level, n_levels - 1))
        current[level] = r["_val"]
        for i in range(level + 1, n_levels):
            current[i] = ""
        prev_level = level

        rows.append({
            "工事名":      r["工事名"],
            "工種":        current[0],
            "種別":        current[1],
            "細別":        current[2],
            "名称":        current[3],
            "規格":        r["規格"],
            "単位":        r["単位"],
            "数量（今回）": r["数量（今回）"],
        })

    return pd.DataFrame(rows, columns=OUT_COLS)


# ---------------------------------------------------------------------------
# メイン抽出関数
# ---------------------------------------------------------------------------

def extract_suryo_full(pdf_path: str) -> pd.DataFrame:
    """
    数量総括表PDFから全データを抽出してDataFrameで返す。

    x座標から工種/種別/細別/名称の4階層を判定する。

    Returns:
        列: 工事名, 工種, 種別, 細別, 名称, 規格, 単位, 数量（今回）
    """
    all_records: list[dict] = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            all_records.extend(_extract_page(page))

    if not all_records:
        return pd.DataFrame(columns=OUT_COLS)

    return _build_hierarchy(all_records)


# ---------------------------------------------------------------------------
# Excel書き出し
# ---------------------------------------------------------------------------

def write_suryo_excel(df: pd.DataFrame, output_path: Optional[str] = None) -> bytes:
    """
    DataFrameをExcelに書き出す。

    Args:
        df:          extract_suryo_full() の戻り値
        output_path: 保存先。None の場合はバイト列を返す。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "数量総括表"

    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_font  = Font(bold=True, color=HEADER_FG, size=10)
    header_fill  = PatternFill("solid", fgColor=HEADER_BG)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_align   = Alignment(vertical="top", wrap_text=True)

    # ── ヘッダー行 ──────────────────────────────────────────
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = border
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # ── データ行 ─────────────────────────────────────────────
    # 工種が変わるタイミングで薄い背景色を入れる
    section_fill = PatternFill("solid", fgColor=SECTION_BG)
    prev_kojyo   = None

    for row_idx, (_, row) in enumerate(df.iterrows(), start=2):
        current_kojyo = row.get("工種", "")
        use_fill = current_kojyo and current_kojyo != prev_kojyo

        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=str(value) if value else "")
            cell.alignment = data_align
            cell.border    = border
            if use_fill:
                cell.fill = section_fill

        if use_fill:
            prev_kojyo = current_kojyo

    # ── 列幅の自動調整 ─────────────────────────────────────
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = len(str(col_name))
        for value in df.iloc[:, col_idx - 1]:
            cell_max = max((len(line) for line in str(value).split("\n")), default=0)
            max_len  = max(max_len, cell_max)
        width = min(max(max_len + 2, MIN_WIDTH), MAX_WIDTH)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    if output_path:
        wb.save(output_path)
        return None
    else:
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(f"使い方: python {Path(__file__).name} <数量総括表.pdf> [出力先.xlsx]")
        sys.exit(1)

    pdf_path = sys.argv[1]
    out_path = sys.argv[2] if len(sys.argv) >= 3 else str(Path(pdf_path).with_suffix(".xlsx"))

    print(f"抽出中: {pdf_path}")
    df = extract_suryo_full(pdf_path)
    print(f"  → {len(df)} 行取得")
    print(df.head(15).to_string(max_colwidth=30))

    print(f"\nExcel生成中: {out_path}")
    write_suryo_excel(df, output_path=out_path)
    print(f"保存完了: {out_path}")
